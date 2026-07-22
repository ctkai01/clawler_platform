from __future__ import annotations

import csv
import io
import os
import re
from datetime import date, datetime, timedelta, timezone

import httpx
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from platform_app.api.deps import get_current_user, require_roles
from platform_app.api.schemas import (
    AccordionCategoryCounts,
    AccordionSentimentCounts,
    DocumentCommentOut,
    DocumentDetailOut,
    DocumentListResponse,
    EngagementGrowthPoint,
    EntityNetworkResponse,
    MonitoringOverview,
    OrgEntitySelectRequest,
    OrgEntitySelection,
    OrgKeywordSelection,
    RelatedDocumentItem,
    SourceCreate,
    SourceImportResult,
    SourceOut,
    SourceUpdate,
    SystemStats,
    TopicOut,
)
from platform_app.db.pool import get_pool
from platform_app.pipeline.settings import VALID_MODES, get_classify_mode, set_classify_mode

VALID_PLATFORM_TYPES = {"facebook_group", "facebook_page", "facebook_profile", "forum", "news"}


def _normalize_url(url: str) -> str:
    """Collapses cosmetic URL variants (protocol, www., trailing slash) so
    duplicate-detection catches the same page imported under different
    forms — e.g. a CSV re-import using bare 'facebook.com/x' must be
    recognized against an existing 'https://www.facebook.com/x/' row."""
    return re.sub(r"^https?://(www\.)?", "", url.strip().rstrip("/")).lower()


def _tracked_entity_condition(alias: str) -> str:
    """Restricts an entity display to canonical_names the caller's org has
    actually selected to track (organization_entities). entity_match.py tags
    every document against the FULL global gazetteer regardless of org
    selection — organization_entities is otherwise just a UI preference —
    so without this filter, org-facing entity displays (badges, network
    graphs, related-post matching) leak whichever gazetteer rows happen to
    be active for ANY industry, not just what this org opted into."""
    return f"{alias}.canonical_name IN (SELECT canonical_name FROM organization_entities WHERE organization_id = %s)"

router = APIRouter(prefix="/org", tags=["org"], dependencies=[Depends(require_roles("org_main", "org_sub"))])

# Configurator-only mutations: org_main always qualifies; org_sub needs the
# 'configurator' functional_role. Kept as its own dependency (not just a
# frontend guard) per the "enforce data-scope server-side" note in the
# architecture doc — an org_sub with report_viewer must not be able to
# write, even by calling the API directly.
def _require_configurator(user: dict = Depends(get_current_user)) -> dict:
    if user["role"] == "org_main":
        return user
    if user["role"] == "org_sub" and user["functional_role"] == "configurator":
        return user
    raise HTTPException(status.HTTP_403_FORBIDDEN, "Cần quyền configurator")


def _owns_target(conn, user: dict, target_id: int) -> bool:
    """A target is writable by this user iff it belongs to their org, and
    (for org_sub) is in their explicitly-granted access list."""
    row = conn.execute(
        "SELECT organization_id FROM crawl_targets WHERE id = %s", (target_id,)
    ).fetchone()
    if row is None or row["organization_id"] != user["organization_id"]:
        return False
    if user["role"] == "org_sub" and target_id not in (user["accessible_target_ids"] or []):
        return False
    return True


# ---------------------------------------------------------------------------
# Sources (crawl_targets scoped to the caller's organization)
# ---------------------------------------------------------------------------


@router.get("/sources", response_model=list[SourceOut])
def list_sources(user: dict = Depends(get_current_user)) -> list[dict]:
    with get_pool().connection() as conn:
        if user["role"] == "org_sub":
            ids = user["accessible_target_ids"] or []
            if not ids:
                return []
            return conn.execute(
                "SELECT * FROM crawl_targets WHERE organization_id = %s AND id = ANY(%s) ORDER BY created_at DESC",
                (user["organization_id"], ids),
            ).fetchall()
        return conn.execute(
            "SELECT * FROM crawl_targets WHERE organization_id = %s ORDER BY created_at DESC",
            (user["organization_id"],),
        ).fetchall()


@router.post("/sources", response_model=SourceOut, status_code=status.HTTP_201_CREATED)
def create_source(body: SourceCreate, user: dict = Depends(_require_configurator)) -> dict:
    with get_pool().connection() as conn:
        existing = conn.execute(
            """
            SELECT id FROM crawl_targets
            WHERE platform_type = %s AND organization_id = %s
            AND regexp_replace(regexp_replace(lower(url), '^https?://(www\\.)?', ''), '/$', '') = %s
            """,
            (body.platform_type, user["organization_id"], _normalize_url(body.url)),
        ).fetchone()
        if existing is not None:
            raise HTTPException(status.HTTP_409_CONFLICT, "Nguồn crawl đã tồn tại")
        return conn.execute(
            """
            INSERT INTO crawl_targets (platform_type, url, display_name, crawl_interval_sec, organization_id)
            VALUES (%s, %s, %s, %s, %s)
            RETURNING *
            """,
            (body.platform_type, body.url, body.display_name, body.crawl_interval_sec, user["organization_id"]),
        ).fetchone()


@router.post("/sources/import", response_model=SourceImportResult)
def import_sources(file: UploadFile = File(...), user: dict = Depends(_require_configurator)) -> dict:
    """Bulk-add sources from a CSV with columns: platform_type,url,display_name
    (display_name optional). Bad/duplicate rows are skipped (not fatal) so
    one typo doesn't block the rest of the file — reasons are returned in
    `errors` keyed by row number so the user can fix and re-upload just
    those rows."""
    raw = file.file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(raw))

    if reader.fieldnames is None or "platform_type" not in reader.fieldnames or "url" not in reader.fieldnames:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "File CSV cần có cột 'platform_type' và 'url' (tuỳ chọn: 'display_name').",
        )

    inserted = 0
    errors: list[str] = []
    total_rows = 0

    with get_pool().connection() as conn:
        for i, row in enumerate(reader, start=2):  # start=2: row 1 is the header
            total_rows += 1
            platform_type = (row.get("platform_type") or "").strip()
            url = (row.get("url") or "").strip()
            display_name = (row.get("display_name") or "").strip() or None

            if platform_type not in VALID_PLATFORM_TYPES:
                errors.append(f"Dòng {i}: platform_type '{platform_type}' không hợp lệ")
                continue
            if not url:
                errors.append(f"Dòng {i}: thiếu url")
                continue

            dup = conn.execute(
                """
                SELECT 1 FROM crawl_targets
                WHERE organization_id = %s AND platform_type = %s
                AND regexp_replace(regexp_replace(lower(url), '^https?://(www\\.)?', ''), '/$', '') = %s
                """,
                (user["organization_id"], platform_type, _normalize_url(url)),
            ).fetchone()
            if dup is not None:
                errors.append(f"Dòng {i}: nguồn đã tồn tại, bỏ qua")
                continue

            result = conn.execute(
                """
                INSERT INTO crawl_targets (platform_type, url, display_name, organization_id)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (organization_id, platform_type, url) WHERE organization_id IS NOT NULL DO NOTHING
                RETURNING id
                """,
                (platform_type, url, display_name, user["organization_id"]),
            ).fetchone()
            if result is not None:
                inserted += 1
            else:
                errors.append(f"Dòng {i}: nguồn đã tồn tại, bỏ qua")

    return {
        "total_rows": total_rows,
        "inserted": inserted,
        "skipped": total_rows - inserted,
        "errors": errors,
    }


@router.patch("/sources/{target_id}", response_model=SourceOut)
def update_source(target_id: int, body: SourceUpdate, user: dict = Depends(_require_configurator)) -> dict:
    with get_pool().connection() as conn:
        if not _owns_target(conn, user, target_id):
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy nguồn crawl")
        return conn.execute(
            "UPDATE crawl_targets SET display_name = %s, updated_at = now() WHERE id = %s RETURNING *",
            (body.display_name, target_id),
        ).fetchone()


@router.delete("/sources/{target_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_source(target_id: int, user: dict = Depends(_require_configurator)) -> None:
    with get_pool().connection() as conn:
        if not _owns_target(conn, user, target_id):
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy nguồn crawl")
        conn.execute("DELETE FROM crawl_targets WHERE id = %s", (target_id,))


# ---------------------------------------------------------------------------
# Monitoring — crawl source health, Airflow DAG run history, document throughput
# ---------------------------------------------------------------------------

# The 5 DAGs behind crawling+classification, in the order shown on the
# monitoring page — same list as platform_app.dashboard.app.CRAWL_DAGS.
_MONITORED_DAGS = [
    "facebook_groups_crawl",
    "facebook_pages_crawl",
    "facebook_profiles_crawl",
    "forums_crawl",
    "news_crawl",
    "content_pipeline",
]
_AIRFLOW_API_BASE = os.environ.get("AIRFLOW_API_BASE", "http://airflow-webserver:8080")
_AIRFLOW_API_USER = os.environ.get("AIRFLOW_API_USER", "admin")
_AIRFLOW_API_PASSWORD = os.environ.get("AIRFLOW_API_PASSWORD", "admin")


def _accessible_target_filter(user: dict) -> tuple[str, list]:
    """WHERE-clause fragment + params scoping crawl_targets/documents to the
    caller's org, additionally restricted to accessible_target_ids for
    org_sub — mirrors the ad-hoc checks used throughout this router."""
    if user["role"] == "org_sub":
        return "ct.organization_id = %s AND ct.id = ANY(%s)", [user["organization_id"], user["accessible_target_ids"] or []]
    return "ct.organization_id = %s", [user["organization_id"]]


def _fetch_recent_dag_runs() -> tuple[list[dict], bool]:
    """Last 5 runs per monitored DAG via Airflow's REST API (no direct DB
    access to the separate `airflow` Postgres database). Returns
    (runs, unreachable) — on any Airflow API failure, returns an empty list
    with unreachable=True rather than failing the whole monitoring page."""
    runs: list[dict] = []
    try:
        with httpx.Client(auth=(_AIRFLOW_API_USER, _AIRFLOW_API_PASSWORD), timeout=10.0) as client:
            for dag_id in _MONITORED_DAGS:
                resp = client.get(
                    f"{_AIRFLOW_API_BASE}/api/v1/dags/{dag_id}/dagRuns",
                    params={"order_by": "-execution_date", "limit": 5},
                )
                resp.raise_for_status()
                for run in resp.json().get("dag_runs", []):
                    start = run.get("start_date")
                    end = run.get("end_date")
                    duration_sec = None
                    if start and end:
                        duration_sec = (
                            datetime.fromisoformat(end) - datetime.fromisoformat(start)
                        ).total_seconds()
                    runs.append(
                        {
                            "dag_id": dag_id,
                            "run_id": run.get("dag_run_id"),
                            "state": run.get("state"),
                            "execution_date": run.get("execution_date"),
                            "start_date": start,
                            "end_date": end,
                            "duration_sec": duration_sec,
                        }
                    )
    except httpx.HTTPError:
        return [], True
    return runs, False


def _get_system_stats() -> dict | None:
    """CPU/RAM/disk of the host this container runs on. Reads /proc directly
    (via psutil) rather than the docker socket — no extra mount/permissions
    needed, and on a single-VPS deployment (no cgroup memory/cpu limits set
    on the api container) /proc reflects the host, not just this container.
    Returns None if psutil is unavailable or reading fails, so the rest of
    the monitoring page still renders."""
    try:
        import psutil

        mem = psutil.virtual_memory()
        disk = psutil.disk_usage("/")
        return {
            "cpu_percent": psutil.cpu_percent(interval=0.2),
            "mem_percent": mem.percent,
            "mem_used_gb": mem.used / 1_073_741_824,
            "mem_total_gb": mem.total / 1_073_741_824,
            "disk_percent": disk.percent,
            "disk_used_gb": disk.used / 1_073_741_824,
            "disk_total_gb": disk.total / 1_073_741_824,
            "load_avg_1m": os.getloadavg()[0],
        }
    except Exception:
        return None


@router.get("/monitoring/system", response_model=SystemStats)
def get_monitoring_system(_user: dict = Depends(get_current_user)) -> dict:
    """Split out from /monitoring/overview so the frontend can poll this one
    much more frequently (near-realtime CPU/RAM) without re-running the
    overview's DB queries + 5 sequential Airflow API calls every tick."""
    stats = _get_system_stats()
    if stats is None:
        raise HTTPException(status.HTTP_503_SERVICE_UNAVAILABLE, "Không đọc được số liệu CPU/RAM")
    return stats


@router.get("/monitoring/overview", response_model=MonitoringOverview)
def get_monitoring_overview(user: dict = Depends(get_current_user)) -> dict:
    where_clause, params = _accessible_target_filter(user)

    with get_pool().connection() as conn:
        if user["role"] == "org_sub" and not (user["accessible_target_ids"] or []):
            sources_by_status: list[dict] = []
            failing_sources: list[dict] = []
            crawled_sources: list[dict] = []
            document_throughput: list[dict] = []
            document_throughput_matched: list[dict] = []
            document_publish_timeline: list[dict] = []
            document_publish_timeline_matched: list[dict] = []
            recent_documents: list[dict] = []
        else:
            sources_by_status = conn.execute(
                f"""
                SELECT ct.platform_type, COALESCE(ct.last_status, 'chua_crawl') AS status, count(*) AS count
                FROM crawl_targets ct
                WHERE {where_clause}
                GROUP BY 1, 2
                ORDER BY 1, 2
                """,
                params,
            ).fetchall()

            failing_sources = conn.execute(
                f"""
                SELECT ct.id, ct.platform_type, ct.display_name, ct.url, ct.last_status,
                       ct.last_error, ct.consecutive_failures, ct.last_crawled_at, ct.fb_session_key
                FROM crawl_targets ct
                WHERE {where_clause} AND ct.last_status IN ('error', 'session_expired')
                ORDER BY ct.consecutive_failures DESC, ct.last_crawled_at DESC NULLS LAST
                LIMIT 500
                """,
                params,
            ).fetchall()

            crawled_sources = conn.execute(
                f"""
                SELECT ct.id, ct.platform_type, ct.display_name, ct.url, ct.last_status, ct.last_crawled_at,
                       (SELECT count(*) FROM documents d WHERE d.target_id = ct.id) AS document_count
                FROM crawl_targets ct
                WHERE {where_clause} AND ct.last_status = 'ok'
                ORDER BY ct.last_crawled_at DESC NULLS LAST
                LIMIT 200
                """,
                params,
            ).fetchall()

            document_throughput = conn.execute(
                f"""
                SELECT date(d.first_seen_at AT TIME ZONE 'Asia/Ho_Chi_Minh') AS day, d.platform_type, count(*) AS count
                FROM documents d
                JOIN crawl_targets ct ON ct.id = d.target_id
                WHERE {where_clause} AND d.first_seen_at >= now() - interval '14 days'
                GROUP BY 1, 2
                ORDER BY 1
                """,
                params,
            ).fetchall()

            # Same chart, but scoped to documents that actually passed the
            # keyword filter (keyword_status='matched' — same concept
            # run_keyword_filter/brand_focus use everywhere else in the
            # platform) — the raw chart above counts every crawled document
            # regardless of relevance, which can look inflated next to how
            # few are ever worth reporting on.
            document_throughput_matched = conn.execute(
                f"""
                SELECT date(d.first_seen_at AT TIME ZONE 'Asia/Ho_Chi_Minh') AS day, d.platform_type, count(*) AS count
                FROM documents d
                JOIN crawl_targets ct ON ct.id = d.target_id
                WHERE {where_clause} AND d.first_seen_at >= now() - interval '14 days'
                  AND d.keyword_status = 'matched'
                GROUP BY 1, 2
                ORDER BY 1
                """,
                params,
            ).fetchall()

            # Same shape again, but grouped by published_at (the post's own
            # date) instead of first_seen_at (when the crawler found it) —
            # a document crawled today about a post from a week ago shows up
            # under today in the two charts above, but under its real
            # publish date here. Needed to actually see "how much of 13-19/7
            # did we capture" instead of "how much did we crawl today".
            document_publish_timeline = conn.execute(
                f"""
                SELECT date(d.published_at AT TIME ZONE 'Asia/Ho_Chi_Minh') AS day, d.platform_type, count(*) AS count
                FROM documents d
                JOIN crawl_targets ct ON ct.id = d.target_id
                WHERE {where_clause} AND d.published_at >= now() - interval '14 days'
                GROUP BY 1, 2
                ORDER BY 1
                """,
                params,
            ).fetchall()

            document_publish_timeline_matched = conn.execute(
                f"""
                SELECT date(d.published_at AT TIME ZONE 'Asia/Ho_Chi_Minh') AS day, d.platform_type, count(*) AS count
                FROM documents d
                JOIN crawl_targets ct ON ct.id = d.target_id
                WHERE {where_clause} AND d.published_at >= now() - interval '14 days'
                  AND d.keyword_status = 'matched'
                GROUP BY 1, 2
                ORDER BY 1
                """,
                params,
            ).fetchall()

            recent_documents = conn.execute(
                f"""
                SELECT d.id, d.platform_type, d.topic, d.url, ct.display_name AS target_name,
                       d.first_seen_at, d.published_at
                FROM documents d
                JOIN crawl_targets ct ON ct.id = d.target_id
                WHERE {where_clause}
                ORDER BY d.first_seen_at DESC
                LIMIT 50
                """,
                params,
            ).fetchall()

    dag_runs, airflow_unreachable = _fetch_recent_dag_runs()

    return {
        "sources_by_status": sources_by_status,
        "failing_sources": failing_sources,
        "crawled_sources": crawled_sources,
        "document_throughput": document_throughput,
        "document_throughput_matched": document_throughput_matched,
        "document_publish_timeline": document_publish_timeline,
        "document_publish_timeline_matched": document_publish_timeline_matched,
        "dag_runs": dag_runs,
        "recent_documents": recent_documents,
        "airflow_unreachable": airflow_unreachable,
    }


# ---------------------------------------------------------------------------
# Entity / keyword tracking selection
# ---------------------------------------------------------------------------


@router.get("/entities", response_model=list[OrgEntitySelection])
def list_org_entities(user: dict = Depends(get_current_user)) -> list[dict]:
    with get_pool().connection() as conn:
        return conn.execute(
            """
            SELECT eg.canonical_name AS canonical_name,
                   eg.industry_code AS industry_code,
                   (oe.organization_id IS NOT NULL) AS is_selected
            FROM (
                SELECT canonical_name, MAX(industry_code) AS industry_code
                FROM entity_gazetteer WHERE is_active = true
                GROUP BY canonical_name
            ) eg
            LEFT JOIN organization_entities oe
                ON oe.canonical_name = eg.canonical_name AND oe.organization_id = %s
            ORDER BY eg.canonical_name
            """,
            (user["organization_id"],),
        ).fetchall()


@router.get("/topics", response_model=list[TopicOut])
def list_org_topics(user: dict = Depends(get_current_user)) -> list[dict]:
    """Read-only view of the chủ đề/keyword taxonomy admin has set up for
    this org (see /admin/topics) — orgs don't manage this themselves, but
    need to be able to see what's already configured."""
    with get_pool().connection() as conn:
        topics = conn.execute(
            "SELECT id, name FROM organization_topics WHERE organization_id = %s ORDER BY name",
            (user["organization_id"],),
        ).fetchall()
        keywords = conn.execute(
            """
            SELECT otk.id, otk.topic_id, otk.keyword
            FROM organization_topic_keywords otk
            JOIN organization_topics ot ON ot.id = otk.topic_id
            WHERE ot.organization_id = %s
            ORDER BY otk.keyword
            """,
            (user["organization_id"],),
        ).fetchall()
    keywords_by_topic: dict[int, list[dict]] = {}
    for kw in keywords:
        keywords_by_topic.setdefault(kw["topic_id"], []).append({"id": kw["id"], "keyword": kw["keyword"]})
    return [{"id": t["id"], "name": t["name"], "keywords": keywords_by_topic.get(t["id"], [])} for t in topics]


@router.post("/entities/select", status_code=status.HTTP_204_NO_CONTENT)
def select_entity(body: OrgEntitySelectRequest, user: dict = Depends(_require_configurator)) -> None:
    with get_pool().connection() as conn:
        conn.execute(
            """
            INSERT INTO organization_entities (organization_id, canonical_name)
            VALUES (%s, %s) ON CONFLICT DO NOTHING
            """,
            (user["organization_id"], body.canonical_name),
        )


@router.post("/entities/deselect", status_code=status.HTTP_204_NO_CONTENT)
def deselect_entity(body: OrgEntitySelectRequest, user: dict = Depends(_require_configurator)) -> None:
    with get_pool().connection() as conn:
        conn.execute(
            "DELETE FROM organization_entities WHERE organization_id = %s AND canonical_name = %s",
            (user["organization_id"], body.canonical_name),
        )


@router.get("/keywords", response_model=list[OrgKeywordSelection])
def list_org_keywords(user: dict = Depends(get_current_user)) -> list[dict]:
    with get_pool().connection() as conn:
        return conn.execute(
            """
            SELECT kc.id AS keyword_id, kc.category, kc.term,
                   (ok.organization_id IS NOT NULL) AS is_selected
            FROM keywords_catalog kc
            LEFT JOIN organization_keywords ok
                ON ok.keyword_id = kc.id AND ok.organization_id = %s
            WHERE kc.is_active = true
            ORDER BY kc.category, kc.term
            """,
            (user["organization_id"],),
        ).fetchall()


@router.post("/keywords/{keyword_id}/select", status_code=status.HTTP_204_NO_CONTENT)
def select_keyword(keyword_id: int, user: dict = Depends(_require_configurator)) -> None:
    with get_pool().connection() as conn:
        conn.execute(
            "INSERT INTO organization_keywords (organization_id, keyword_id) VALUES (%s, %s) ON CONFLICT DO NOTHING",
            (user["organization_id"], keyword_id),
        )


@router.delete("/keywords/{keyword_id}/select", status_code=status.HTTP_204_NO_CONTENT)
def deselect_keyword(keyword_id: int, user: dict = Depends(_require_configurator)) -> None:
    with get_pool().connection() as conn:
        conn.execute(
            "DELETE FROM organization_keywords WHERE organization_id = %s AND keyword_id = %s",
            (user["organization_id"], keyword_id),
        )


# ---------------------------------------------------------------------------
# Report — same shape of numbers as the internal /report dashboard page,
# but scoped to the caller's organization (and further to their granted
# targets if they're an org_sub) rather than a free-text entity filter.
# ---------------------------------------------------------------------------


_REPORT_CHANNEL_PREFIX = {
    "facebook_group": "FB Group",
    "facebook_page": "FB Page",
    "facebook_profile": "FB Profile",
    "forum": "Forum",
    "news": "News",
}

_EMPTY_REPORT = {
    "total_posts": 0,
    "total_comments": 0,
    "total_reactions": 0,
    "total_shares": 0,
    "sentiment_positive": 0,
    "sentiment_negative": 0,
    "sentiment_neutral": 0,
    "topic_detail": [],
    "keyword_topic_detail": [],
    "keyword_topic_sentiment": [],
    "topics": [],
    "topic_positive_counts": [],
    "topic_neutral_counts": [],
    "topic_negative_counts": [],
    "negative_count": 0,
    "positive_count": 0,
    "negative_posts": [],
    "positive_posts": [],
}


def _report_scope(
    user: dict, days: int, entity: str | None, brand_focus: str = "own"
) -> tuple[list[str], list, str, list] | None:
    """Shared org/date-range/entity scoping for /report and /report/posts.
    Returns None when an org_sub has no granted targets (caller should
    short-circuit to an empty result) — otherwise (conditions, params,
    entity_clause, entity_params)."""
    period_end = datetime.now(timezone.utc)
    period_start = period_end - timedelta(days=days)
    return _report_scope_between(user, period_start, period_end, entity, brand_focus)


def _report_scope_between(
    user: dict, period_start: datetime, period_end: datetime, entity: str | None, brand_focus: str = "own"
) -> tuple[list[str], list, str, list] | None:
    """Same as _report_scope but takes an explicit [period_start, period_end)
    window instead of a rolling `days` count — used by the Word daily report
    (fixed 08:00-to-08:00 window) instead of the dashboard's flexible
    "last N days" picker.

    `brand_focus` ('own' | 'competitor') scopes to documents keyword_filter
    tagged as being about this org's own brand vs. only a competitor — the
    condition lives here (not per-query) so every report query built from
    the returned `conditions` inherits it for free."""
    conditions = ["ct.organization_id = %s", "d.brand_focus = %s"]
    params: list = [user["organization_id"], brand_focus]
    if user["role"] == "org_sub":
        ids = user["accessible_target_ids"] or []
        if not ids:
            return None
        conditions.append("d.target_id = ANY(%s)")
        params.append(ids)
    conditions.append("d.published_at >= %s AND d.published_at <= %s")
    params.extend([period_start, period_end])

    entity_clause = ""
    entity_params: list = []
    if entity and entity.strip():
        entity_clause = (
            f"AND EXISTS (SELECT 1 FROM document_entities de WHERE de.document_id = d.id "
            f"AND de.concept_id != '__none__' AND {_tracked_entity_condition('de')} AND de.canonical_name ILIKE %s)"
        )
        entity_params = [user["organization_id"], f"%{entity.strip()}%"]

    return conditions, params, entity_clause, entity_params


def _report_post_row(row: dict) -> dict:
    title = row["topic"] or (row["content"][:80] + "…" if len(row["content"]) > 80 else row["content"])
    channel_prefix = _REPORT_CHANNEL_PREFIX.get(row["platform_type"], row["platform_type"])
    channel_label = f"{channel_prefix}: {row['target_name']}"
    return {
        "id": row["id"],
        "title": title,
        "url": row["url"],
        "channel_label": channel_label,
        "author": row["author"],
        "engagement_total": row["engagement_total"],
    }


@router.get("/report/posts")
def report_posts(
    sentiment: str = Query(..., pattern="^(positive|negative)$"),
    days: int = Query(default=7, ge=1, le=365),
    entity: str | None = Query(default=None),
    scope: str = Query(default="own", pattern="^(own|competitor)$"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=10, ge=1, le=100),
    user: dict = Depends(get_current_user),
) -> dict:
    """Paginated version of /report's negative_posts/positive_posts (which
    only ever return a top-5 preview) — powers the "xem thêm" list."""
    report_scope = _report_scope(user, days, entity, scope)
    if report_scope is None:
        return {"items": [], "total": 0}
    conditions, params, entity_clause, entity_params = report_scope

    post_conditions = [*conditions, "d.classification_sentiment = %s"]
    post_params = [*params, sentiment]
    where = " AND ".join(post_conditions)
    full_params = [*post_params, *entity_params]

    with get_pool().connection() as conn:
        total = conn.execute(
            f"""
            SELECT COUNT(*) AS n
            FROM documents d
            JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE {where} {entity_clause}
            """,
            full_params,
        ).fetchone()["n"]

        rows = conn.execute(
            f"""
            SELECT d.id, d.topic, d.content, d.url, d.author, d.platform_type,
                   ct.display_name AS target_name,
                   (COALESCE(d.reaction_count, 0) + COALESCE(d.comment_count, 0)) AS engagement_total
            FROM documents d
            JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE {where} {entity_clause}
            ORDER BY engagement_total DESC, d.published_at DESC
            LIMIT %s OFFSET %s
            """,
            [*full_params, page_size, (page - 1) * page_size],
        ).fetchall()

    return {"items": [_report_post_row(r) for r in rows], "total": total}


def _build_report(user: dict, days: int, entity: str | None, brand_focus: str = "own") -> dict | None:
    """Shared by /report and /report/export — returns None when an org_sub
    has no granted targets (caller should fall back to _EMPTY_REPORT)."""
    scope = _report_scope(user, days, entity, brand_focus)
    if scope is None:
        return None
    return _build_report_from_scope(user, scope, entity)


def _build_report_between(
    user: dict, period_start: datetime, period_end: datetime, entity: str | None
) -> dict | None:
    """Same as _build_report but for an explicit [period_start, period_end)
    window — used by the Word daily report's fixed 08:00-to-08:00 window."""
    scope = _report_scope_between(user, period_start, period_end, entity)
    if scope is None:
        return None
    return _build_report_from_scope(user, scope, entity)


def _build_report_from_scope(
    user: dict, scope: tuple[list[str], list, str, list], entity: str | None
) -> dict:
    conditions, params, entity_clause, entity_params = scope

    where_clause = " AND ".join(conditions)
    full_params = [*params, *entity_params]

    with get_pool().connection() as conn:
        kpis = conn.execute(
            f"""
            SELECT
                COUNT(*) AS total_posts,
                COALESCE(SUM(d.comment_count), 0) AS total_comments,
                COALESCE(SUM(d.reaction_count), 0) AS total_reactions,
                COALESCE(SUM(d.share_count), 0) AS total_shares
            FROM documents d
            JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE {where_clause} AND d.keyword_status = 'matched' {entity_clause}
            """,
            full_params,
        ).fetchone()

        sentiment_rows = conn.execute(
            f"""
            SELECT d.classification_sentiment AS sentiment, COUNT(*) AS count
            FROM documents d
            JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE {where_clause}
              AND d.classification_status = 'completed' AND d.classification_sentiment IS NOT NULL
            {entity_clause}
            GROUP BY d.classification_sentiment
            """,
            full_params,
        ).fetchall()

        topic_conditions = [*conditions, _tracked_entity_condition("de")]
        topic_params = [*params, user["organization_id"]]
        if entity and entity.strip():
            topic_conditions.append("de.canonical_name ILIKE %s")
            topic_params.append(f"%{entity.strip()}%")
        topic_where = " AND ".join(topic_conditions)

        topic_detail = conn.execute(
            f"""
            SELECT de.canonical_name AS topic,
                   COUNT(DISTINCT d.id) AS posts,
                   COALESCE(SUM(d.comment_count), 0) AS comments,
                   COALESCE(SUM(d.reaction_count) + SUM(d.comment_count), 0) AS total_engagement
            FROM documents d
            JOIN crawl_targets ct ON ct.id = d.target_id
            JOIN document_entities de ON de.document_id = d.id AND de.concept_id != '__none__'
            WHERE {topic_where} AND d.keyword_status = 'matched'
            GROUP BY de.canonical_name
            ORDER BY posts DESC
            LIMIT 15
            """,
            topic_params,
        ).fetchall()

        topics = [row["topic"] for row in topic_detail]
        topic_sentiment_map = {t: {"positive": 0, "negative": 0, "neutral": 0} for t in topics}
        if topics:
            topic_sentiment_rows = conn.execute(
                f"""
                SELECT de.canonical_name AS topic, d.classification_sentiment AS sentiment,
                       COUNT(DISTINCT d.id) AS count
                FROM documents d
                JOIN crawl_targets ct ON ct.id = d.target_id
                JOIN document_entities de ON de.document_id = d.id AND de.concept_id != '__none__'
                WHERE {where_clause} AND de.canonical_name = ANY(%s)
                  AND d.keyword_status = 'matched'
                  AND d.classification_status = 'completed' AND d.classification_sentiment IS NOT NULL
                GROUP BY de.canonical_name, d.classification_sentiment
                """,
                [*params, topics],
            ).fetchall()
            for row in topic_sentiment_rows:
                if row["topic"] in topic_sentiment_map:
                    topic_sentiment_map[row["topic"]][row["sentiment"]] = row["count"]

        keyword_topic_detail = conn.execute(
            f"""
            SELECT COALESCE(ot.name, 'KHÁC') AS topic,
                   COUNT(*) AS posts,
                   COALESCE(SUM(d.comment_count), 0) AS comments,
                   COALESCE(SUM(d.reaction_count) + SUM(d.comment_count), 0) AS total_engagement
            FROM documents d
            JOIN crawl_targets ct ON ct.id = d.target_id
            LEFT JOIN organization_topics ot ON ot.id = d.topic_tag_id
            WHERE {where_clause} AND d.keyword_status = 'matched' {entity_clause}
            GROUP BY ot.name
            ORDER BY posts DESC
            """,
            full_params,
        ).fetchall()

        keyword_topic_sentiment = conn.execute(
            f"""
            SELECT COALESCE(ot.name, 'KHÁC') AS topic,
                   COUNT(*) FILTER (WHERE d.classification_sentiment = 'positive') AS positive,
                   COUNT(*) FILTER (WHERE d.classification_sentiment = 'neutral') AS neutral,
                   COUNT(*) FILTER (WHERE d.classification_sentiment = 'negative') AS negative
            FROM documents d
            JOIN crawl_targets ct ON ct.id = d.target_id
            LEFT JOIN organization_topics ot ON ot.id = d.topic_tag_id
            WHERE {where_clause} {entity_clause}
              AND d.classification_status = 'completed' AND d.classification_sentiment IS NOT NULL
            GROUP BY ot.name
            ORDER BY COUNT(*) DESC
            """,
            full_params,
        ).fetchall()

        def _top_posts(sentiment: str, limit: int = 5) -> list[dict]:
            post_conditions = [*conditions, "d.classification_sentiment = %s"]
            post_params = [*params, sentiment]
            rows = conn.execute(
                f"""
                SELECT d.id, d.topic, d.content, d.url, d.author, d.platform_type,
                       ct.display_name AS target_name,
                       (COALESCE(d.reaction_count, 0) + COALESCE(d.comment_count, 0)) AS engagement_total
                FROM documents d
                JOIN crawl_targets ct ON ct.id = d.target_id
                WHERE {" AND ".join(post_conditions)} {entity_clause}
                ORDER BY engagement_total DESC, d.published_at DESC
                LIMIT %s
                """,
                [*post_params, *entity_params, limit],
            ).fetchall()
            return [_report_post_row(r) for r in rows]

        negative_posts = _top_posts("negative")
        positive_posts = _top_posts("positive")

    sentiment_map = {r["sentiment"]: r["count"] for r in sentiment_rows}
    return {
        **kpis,
        "sentiment_positive": sentiment_map.get("positive", 0),
        "sentiment_negative": sentiment_map.get("negative", 0),
        "sentiment_neutral": sentiment_map.get("neutral", 0),
        "topic_detail": topic_detail,
        "keyword_topic_detail": keyword_topic_detail,
        "keyword_topic_sentiment": keyword_topic_sentiment,
        "topics": topics,
        "topic_positive_counts": [topic_sentiment_map[t]["positive"] for t in topics],
        "topic_neutral_counts": [topic_sentiment_map[t]["neutral"] for t in topics],
        "topic_negative_counts": [topic_sentiment_map[t]["negative"] for t in topics],
        "negative_count": sentiment_map.get("negative", 0),
        "positive_count": sentiment_map.get("positive", 0),
        "negative_posts": negative_posts,
        "positive_posts": positive_posts,
    }


def _negative_channel_split(conn, conditions: list[str], params: list) -> dict[str, int]:
    """Negative-sentiment document count split into news vs. everything else
    (social) — used by the negative-brand weekly report's summary table,
    which needs this split but not the full topic/top-posts machinery of
    _build_report_from_scope."""
    where_clause = " AND ".join(conditions)
    rows = conn.execute(
        f"""
        SELECT (d.platform_type = 'news') AS is_news, COUNT(*) AS count
        FROM documents d
        JOIN crawl_targets ct ON ct.id = d.target_id
        WHERE {where_clause} AND d.classification_status = 'completed' AND d.classification_sentiment = 'negative'
        GROUP BY is_news
        """,
        params,
    ).fetchall()
    result = {"news": 0, "social": 0}
    for row in rows:
        result["news" if row["is_news"] else "social"] = row["count"]
    return result


def _negative_top_topic(conn, conditions: list[str], params: list, *, news: bool) -> tuple[str, int] | None:
    """The single most common organization_topics tag among this channel's
    negative documents, with its count — None if the channel has zero
    negative documents. Same grouping as keyword_topic_detail (org.py above)
    but scoped to one channel + negative sentiment only."""
    where_clause = " AND ".join(conditions)
    channel_condition = "d.platform_type = 'news'" if news else "d.platform_type != 'news'"
    row = conn.execute(
        f"""
        SELECT COALESCE(ot.name, 'KHÁC') AS topic, COUNT(*) AS count
        FROM documents d
        JOIN crawl_targets ct ON ct.id = d.target_id
        LEFT JOIN organization_topics ot ON ot.id = d.topic_tag_id
        WHERE {where_clause} AND {channel_condition}
          AND d.classification_status = 'completed' AND d.classification_sentiment = 'negative'
        GROUP BY ot.name
        ORDER BY count DESC
        LIMIT 1
        """,
        params,
    ).fetchone()
    return (row["topic"], row["count"]) if row else None


def _negative_topic_docs(
    conn, conditions: list[str], params: list, *, news: bool, topic: str, limit: int = 10
) -> list[dict]:
    """Real document excerpts for the top negative topic in one channel —
    fed to the LLM to write the "Nội dung tiêu cực chính" summary sentence,
    same content-excerpt pattern as event_report.py's overview narrative."""
    where_clause = " AND ".join(conditions)
    channel_condition = "d.platform_type = 'news'" if news else "d.platform_type != 'news'"
    if topic == "KHÁC":
        topic_condition = "ot.name IS NULL"
        topic_params = list(params)
    else:
        topic_condition = "ot.name = %s"
        topic_params = [*params, topic]
    rows = conn.execute(
        f"""
        SELECT d.topic, d.content, ct.display_name AS target_name
        FROM documents d
        JOIN crawl_targets ct ON ct.id = d.target_id
        LEFT JOIN organization_topics ot ON ot.id = d.topic_tag_id
        WHERE {where_clause} AND {channel_condition} AND {topic_condition}
          AND d.classification_status = 'completed' AND d.classification_sentiment = 'negative'
        ORDER BY d.published_at DESC
        LIMIT %s
        """,
        [*topic_params, limit],
    ).fetchall()
    return list(rows)


def _brand_scope_condition(user: dict, period_start: datetime, period_end: datetime) -> tuple[list[str], list] | None:
    """Same access-control shape as _report_scope_between (org scoping,
    org_sub target restriction, date window) but WITHOUT brand_focus/entity
    filters — for reports comparing multiple brands symmetrically (own
    brand + competitors as peers) rather than scoping to one org's own
    content. None when an org_sub has no granted targets, same convention
    as _report_scope_between."""
    conditions = ["ct.organization_id = %s"]
    params: list = [user["organization_id"]]
    if user["role"] == "org_sub":
        ids = user["accessible_target_ids"] or []
        if not ids:
            return None
        conditions.append("d.target_id = ANY(%s)")
        params.append(ids)
    conditions.append("d.published_at >= %s AND d.published_at <= %s")
    params.extend([period_start, period_end])
    return conditions, params


def _brand_sentiment_counts(conn, conditions: list[str], params: list, brands: list[str]) -> dict[str, dict[str, int]]:
    """Per-brand sentiment breakdown using the already-computed
    documents.classification_sentiment (no new LLM call needed) — same
    grouping pattern as _build_report_from_scope's topic_sentiment_map
    (above) but scoped to a fixed brand list instead of
    _tracked_entity_condition's org-selected-entities filter.

    Case-insensitive match on canonical_name: organizations.name (used as
    the org's own brand string here) and entity_gazetteer's seeded
    canonical_name can disagree on casing for the same brand (confirmed in
    prod data: organizations.name = "Mobifone", entity_gazetteer =
    "MobiFone") — an exact `=` would silently return zero rows for the
    org's own brand while still matching competitors verbatim."""
    where_clause = " AND ".join(conditions)
    upper_to_brand = {b.upper(): b for b in brands}
    rows = conn.execute(
        f"""
        SELECT de.canonical_name AS brand, d.classification_sentiment AS sentiment, COUNT(DISTINCT d.id) AS count
        FROM documents d
        JOIN crawl_targets ct ON ct.id = d.target_id
        JOIN document_entities de ON de.document_id = d.id AND UPPER(de.canonical_name) = ANY(%s)
        WHERE {where_clause} AND d.classification_status = 'completed' AND d.classification_sentiment IS NOT NULL
        GROUP BY de.canonical_name, d.classification_sentiment
        """,
        [list(upper_to_brand.keys()), *params],
    ).fetchall()
    result = {brand: {"positive": 0, "neutral": 0, "negative": 0} for brand in brands}
    for row in rows:
        brand = upper_to_brand.get(row["brand"].upper())
        if brand:
            result[brand][row["sentiment"]] = row["count"]
    return result


def _brand_channel_breakdown(conn, conditions: list[str], params: list, brand: str) -> dict[str, int]:
    """Document count for one brand, bucketed into the channels this
    platform actually crawls (VALID_PLATFORM_TYPES, above): Facebook
    (facebook_group + facebook_page + facebook_profile), News, Forum.
    Case-insensitive brand match — see _brand_sentiment_counts docstring."""
    where_clause = " AND ".join(conditions)
    rows = conn.execute(
        f"""
        SELECT d.platform_type, COUNT(DISTINCT d.id) AS count
        FROM documents d
        JOIN crawl_targets ct ON ct.id = d.target_id
        JOIN document_entities de ON de.document_id = d.id AND UPPER(de.canonical_name) = UPPER(%s)
        WHERE {where_clause}
        GROUP BY d.platform_type
        """,
        [brand, *params],
    ).fetchall()
    breakdown = {"Facebook": 0, "News": 0, "Forum": 0}
    bucket = {
        "facebook_group": "Facebook",
        "facebook_page": "Facebook",
        "facebook_profile": "Facebook",
        "news": "News",
        "forum": "Forum",
    }
    for row in rows:
        key = bucket.get(row["platform_type"])
        if key:
            breakdown[key] += row["count"]
    return breakdown


def _brand_top_posts(
    conn, conditions: list[str], params: list, *, brand: str, sentiment: str, limit: int = 3
) -> list[dict]:
    """Top posts mentioning one brand with one sentiment, ranked by
    engagement — same shape as _report_post_row's source query but scoped
    by document_entities (any brand) instead of brand_focus (own-brand
    only), and includes `images` for the competitor post gallery (mục 3).
    Case-insensitive brand match — see _brand_sentiment_counts docstring."""
    where_clause = " AND ".join(conditions)
    rows = conn.execute(
        f"""
        SELECT d.id, d.topic, d.content, d.url, d.author, d.platform_type, d.images,
               ct.display_name AS target_name,
               (COALESCE(d.reaction_count, 0) + COALESCE(d.comment_count, 0)) AS engagement_total
        FROM documents d
        JOIN crawl_targets ct ON ct.id = d.target_id
        JOIN document_entities de ON de.document_id = d.id AND UPPER(de.canonical_name) = UPPER(%s)
        WHERE {where_clause} AND d.classification_status = 'completed' AND d.classification_sentiment = %s
        ORDER BY engagement_total DESC, d.published_at DESC
        LIMIT %s
        """,
        [brand, *params, sentiment, limit],
    ).fetchall()
    return list(rows)


@router.get("/report")
def org_report(
    days: int = Query(default=7, ge=1, le=365),
    entity: str | None = Query(default=None),
    scope: str = Query(default="own", pattern="^(own|competitor)$"),
    user: dict = Depends(get_current_user),
) -> dict:
    """Same shape/sections as the internal /report dashboard (KPIs + sentiment
    pie, topic/entity breakdown table, sentiment-by-topic stacked bar,
    top negative/positive posts) but scoped to the caller's organization (and
    further to their granted targets if org_sub), and "topic" rows are
    restricted to entities THIS org tracks (see _tracked_entity_condition) —
    not the internal report's free-text match against the full gazetteer.

    `scope=own` (default) is this org's own brand — the main dashboard.
    `scope=competitor` is content that only mentions a tracked competitor
    (see keyword_filter.py's brand_focus tagging) — a separate "Đối thủ" tab,
    kept out of the main aggregates so competitor buzz doesn't get counted
    as if it were about this org."""
    report = _build_report(user, days, entity, scope)
    return report if report is not None else _EMPTY_REPORT


_EXPORT_POST_CAP = 5000


def _all_report_posts(user: dict, days: int, entity: str | None, sentiment: str, brand_focus: str = "own") -> list[dict]:
    """Every matching post (not just the report's top-5 preview) — used for
    Excel export, which needs the full list rather than a UI-sized sample."""
    scope = _report_scope(user, days, entity, brand_focus)
    if scope is None:
        return []
    return _all_report_posts_from_scope(scope, sentiment)


def _all_report_posts_between(
    user: dict, period_start: datetime, period_end: datetime, entity: str | None, sentiment: str
) -> list[dict]:
    """Same as _all_report_posts but for an explicit [period_start, period_end)
    window — used by the Word daily report."""
    scope = _report_scope_between(user, period_start, period_end, entity)
    if scope is None:
        return []
    return _all_report_posts_from_scope(scope, sentiment)


def _all_report_posts_from_scope(scope: tuple[list[str], list, str, list], sentiment: str) -> list[dict]:
    conditions, params, entity_clause, entity_params = scope
    post_conditions = [*conditions, "d.classification_sentiment = %s"]
    post_params = [*params, sentiment]
    with get_pool().connection() as conn:
        rows = conn.execute(
            f"""
            SELECT d.id, d.topic, d.content, d.url, d.author, d.platform_type,
                   ct.display_name AS target_name,
                   (COALESCE(d.reaction_count, 0) + COALESCE(d.comment_count, 0)) AS engagement_total
            FROM documents d
            JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE {" AND ".join(post_conditions)} {entity_clause}
            ORDER BY engagement_total DESC, d.published_at DESC
            LIMIT %s
            """,
            [*post_params, *entity_params, _EXPORT_POST_CAP],
        ).fetchall()
    return [_report_post_row(r) for r in rows]


def _autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 80)


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def build_report_workbook_bytes(user: dict, days: int, entity: str | None) -> bytes:
    """Builds the same Excel workbook /report/export streams to the browser
    — shared with the daily report-email job (platform_app.pipeline.report_email)
    so both paths produce byte-identical files from one code path."""
    report = _build_report(user, days, entity) or _EMPTY_REPORT
    negative_posts = _all_report_posts(user, days, entity, "negative")
    positive_posts = _all_report_posts(user, days, entity, "positive")
    return _build_workbook_bytes(report, negative_posts, positive_posts)


def build_report_workbook_bytes_between(
    user: dict, period_start: datetime, period_end: datetime, entity: str | None
) -> bytes:
    """Same as build_report_workbook_bytes but for an explicit [period_start,
    period_end) window — used by the automated daily report-email job, which
    anchors to a fixed schedule window (e.g. 8am-to-8am) rather than a rolling
    "now minus N days" window."""
    report = _build_report_between(user, period_start, period_end, entity) or _EMPTY_REPORT
    negative_posts = _all_report_posts_between(user, period_start, period_end, entity, "negative")
    positive_posts = _all_report_posts_between(user, period_start, period_end, entity, "positive")
    return _build_workbook_bytes(report, negative_posts, positive_posts)


def build_daily_word_report_bytes_between(user: dict, org_name: str, report_date: date) -> bytes:
    """Builds the same Word (.docx) daily report /report/export-word streams
    to the browser — shared with the daily report-email job
    (platform_app.pipeline.report_email) so both paths produce byte-identical
    files from one code path. Always uses the fixed 08:00-to-08:00 (Vietnam
    time) window ending on `report_date`."""
    from platform_app.reporting.word_report import daily_window

    period_start, period_end = daily_window(report_date)
    return build_word_report_bytes_between(user, org_name, report_date, period_start, period_end, None)


def build_word_report_bytes_between(
    user: dict,
    org_name: str,
    report_date: date,
    period_start: datetime,
    period_end: datetime,
    entity: str | None,
    *,
    period_label: str | None = None,
) -> bytes:
    """Same Word (.docx) template as build_daily_word_report_bytes_between but
    for an explicit [period_start, period_end) window — used by the manual
    "Gửi email ngay" button (flexible "last N days" picker) and by
    build_monthly_brand_report_bytes (30-day window). `period_label`
    overrides the title's default "NGÀY {report_date}" — needed for the
    monthly report, whose window isn't a single day."""
    from platform_app.reporting.word_report import build_daily_word_report_bytes

    report = _build_report_between(user, period_start, period_end, entity) or _EMPTY_REPORT
    negative_posts = _all_report_posts_between(user, period_start, period_end, entity, "negative")
    positive_posts = _all_report_posts_between(user, period_start, period_end, entity, "positive")
    return build_daily_word_report_bytes(
        org_name=org_name,
        report_date=report_date,
        report=report,
        topic_sentiment_rows=report["keyword_topic_sentiment"],
        negative_posts=negative_posts,
        positive_posts=positive_posts,
        period_label=period_label,
    )


def _build_workbook_bytes(report: dict, negative_posts: list[dict], positive_posts: list[dict]) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Tổng quan"
    ws.append(["Chỉ số", "Giá trị"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.append(["Tổng số tin", report["total_posts"]])
    ws.append(["Bình luận", report["total_comments"]])
    ws.append(["Tổng reaction", report["total_reactions"]])
    ws.append(["Chia sẻ", report["total_shares"]])
    ws.append(["Tích cực", report["sentiment_positive"]])
    ws.append(["Trung tính", report["sentiment_neutral"]])
    ws.append(["Tiêu cực", report["sentiment_negative"]])
    _autosize_columns(ws)

    ws = wb.create_sheet("Theo chủ đề (entity)")
    _write_header(ws, ["Chủ đề", "Bài đăng", "Bình luận", "Tổng số tương tác"])
    for row in report["topic_detail"]:
        ws.append([row["topic"], row["posts"], row["comments"], row["total_engagement"]])
    _autosize_columns(ws)

    ws = wb.create_sheet("Theo chủ đề (từ khóa)")
    _write_header(ws, ["Chủ đề", "Bài đăng", "Bình luận", "Tổng số tương tác"])
    for row in report["keyword_topic_detail"]:
        ws.append([row["topic"], row["posts"], row["comments"], row["total_engagement"]])
    _autosize_columns(ws)

    for sheet_name, posts in (("Tiêu cực", negative_posts), ("Tích cực", positive_posts)):
        ws = wb.create_sheet(sheet_name)
        _write_header(ws, ["Tiêu đề bài đăng", "Kênh", "Người đăng", "Tổng số tương tác", "Link"])
        for p in posts:
            ws.append([p["title"], p["channel_label"], p["author"] or "", p["engagement_total"], p["url"]])
        _autosize_columns(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/report/export")
def export_report(
    days: int = Query(default=7, ge=1, le=365),
    entity: str | None = Query(default=None),
    user: dict = Depends(get_current_user),
) -> StreamingResponse:
    """Excel export of the report currently shown on screen (same days/entity
    filters), with full negative/positive post lists (up to _EXPORT_POST_CAP)
    instead of the UI's top-5 preview."""
    content = build_report_workbook_bytes(user, days, entity)
    filename = f"bao-cao-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/report/export-word")
def export_report_word(
    report_date: date = Query(default_factory=lambda: datetime.now(timezone.utc).date()),
    user: dict = Depends(get_current_user),
) -> StreamingResponse:
    """Word (.docx) daily report matching the org's existing manual report
    template — fixed 08:00-to-08:00 (Vietnam time) window ending on
    `report_date`, not the dashboard's flexible "last N days" picker."""
    content = build_daily_word_report_bytes_between(user, user["organization_name"], report_date)
    filename = f"bao-cao-{report_date.strftime('%Y%m%d')}.docx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _event_sentiment_counts(matches: list[dict]) -> dict[str, int]:
    counts = {"positive": 0, "neutral": 0, "negative": 0}
    for m in matches:
        counts[m["sentiment"]] = counts.get(m["sentiment"], 0) + 1
    return counts


def compute_event_report_data(user: dict, event_key: str, report_date: date) -> dict:
    """All data build_event_word_report_bytes needs, computed once — reused
    by both the Word export and the JSON preview route
    (/report/event/{event_key}/data)."""
    from platform_app.pipeline.event_report import (
        EVENT_DEFINITIONS,
        generate_overview_narrative,
        get_competitor_brands,
        run_event_match,
    )
    from platform_app.reporting.word_report import daily_window

    event = EVENT_DEFINITIONS[event_key]
    org_name = user["organization_name"]
    organization_id = user["organization_id"]
    yesterday_date = report_date - timedelta(days=1)

    today_start, today_end = daily_window(report_date)
    yesterday_start, yesterday_end = daily_window(yesterday_date)

    today_matches = run_event_match(event_key, organization_id, today_start, today_end)
    yesterday_matches = run_event_match(event_key, organization_id, yesterday_start, yesterday_end)

    today_news = [m for m in today_matches if m["platform_type"] == "news"]
    today_social = [m for m in today_matches if m["platform_type"] != "news"]
    yesterday_news = [m for m in yesterday_matches if m["platform_type"] == "news"]
    yesterday_social = [m for m in yesterday_matches if m["platform_type"] != "news"]

    comparison = {
        "yesterday_label": f"{yesterday_date.day}/{yesterday_date.month}/{yesterday_date.year}",
        "today_label": f"{report_date.day}/{report_date.month}/{report_date.year}",
        "news": {
            "yesterday_total": len(yesterday_news),
            "today_total": len(today_news),
            "yesterday_sentiment": _event_sentiment_counts(yesterday_news),
            "today_sentiment": _event_sentiment_counts(today_news),
        },
        "social": {
            "yesterday_total": len(yesterday_social),
            "today_total": len(today_social),
            "yesterday_sentiment": _event_sentiment_counts(yesterday_social),
            "today_sentiment": _event_sentiment_counts(today_social),
        },
    }

    mobifone_news = [m for m in today_news if m["brand"] == org_name]
    competitor_news: dict[str, list[dict]] = {brand: [] for brand in get_competitor_brands(organization_id)}
    for m in today_news:
        if m["brand"] != org_name:
            competitor_news.setdefault(m["brand"], []).append(m)
    social_mobifone = [m for m in today_social if m["brand"] == org_name]

    overview_narrative = generate_overview_narrative(org_name, today_news)

    return {
        "org_name": org_name,
        "event_label": event["label"],
        "report_date": report_date,
        "comparison": comparison,
        "overview_narrative": overview_narrative,
        "mobifone_news": mobifone_news,
        "competitor_news": competitor_news,
        "social_matches": social_mobifone,
    }


def build_event_word_report_bytes(user: dict, event_key: str, report_date: date) -> bytes:
    """Builds the "sự vụ" (topic-event) Word report — e.g. 5G MobiFone vs
    đối thủ — distinct from build_daily_word_report_bytes_between: scoped to
    one topic (not all org content), split báo chí/mạng xã hội, and includes
    an LLM-written overview comparing MobiFone against its competitors."""
    from platform_app.reporting.event_word_report import build_event_daily_word_report_bytes

    data = compute_event_report_data(user, event_key, report_date)
    return build_event_daily_word_report_bytes(**data)


@router.get("/report/event/{event_key}/export-word")
def export_event_report_word(
    event_key: str,
    report_date: date = Query(default_factory=lambda: datetime.now(timezone.utc).date()),
    user: dict = Depends(get_current_user),
) -> StreamingResponse:
    """Word (.docx) daily "sự vụ" (topic-event) report — e.g. 5G MobiFone vs
    đối thủ — fixed 08:00-to-08:00 (Vietnam time) window ending on
    `report_date`, plus the preceding day's window for the comparison table."""
    from platform_app.pipeline.event_report import EVENT_DEFINITIONS

    if event_key not in EVENT_DEFINITIONS:
        raise HTTPException(status.HTTP_404_NOT_FOUND, f"Không tìm thấy sự vụ '{event_key}'")

    content = build_event_word_report_bytes(user, event_key, report_date)
    filename = f"bao-cao-{event_key}-{report_date.strftime('%Y%m%d')}.docx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/report/event/{event_key}/data")
def get_event_report_data(
    event_key: str,
    report_date: date = Query(default_factory=lambda: datetime.now(timezone.utc).date()),
    user: dict = Depends(get_current_user),
) -> dict:
    """JSON preview of the same data export_event_report_word renders to
    .docx — lets the Báo cáo tab show the report on-screen before export."""
    from platform_app.pipeline.event_report import EVENT_DEFINITIONS

    if event_key not in EVENT_DEFINITIONS:
        raise HTTPException(status.HTTP_404_NOT_FOUND, f"Không tìm thấy sự vụ '{event_key}'")

    return compute_event_report_data(user, event_key, report_date)


def _range_label(start_date: date, end_date: date) -> str:
    return f"{start_date.day:02d}/{start_date.month:02d} - {end_date.day:02d}/{end_date.month:02d}"


def compute_event_weekly_report_data(user: dict, event_key: str, report_date: date) -> dict:
    """All data build_event_weekly_word_report_bytes needs, computed once —
    reused by both the Word export and the JSON preview route
    (/report/event/{event_key}/data-weekly)."""
    from platform_app.pipeline.event_report import (
        EVENT_DEFINITIONS,
        generate_overview_narrative,
        get_competitor_brands,
        run_event_match,
    )
    from platform_app.reporting.word_report import weekly_window

    event = EVENT_DEFINITIONS[event_key]
    org_name = user["organization_name"]
    organization_id = user["organization_id"]
    prev_report_date = report_date - timedelta(days=7)

    this_start, this_end = weekly_window(report_date)
    prev_start, prev_end = weekly_window(prev_report_date)

    this_matches = run_event_match(event_key, organization_id, this_start, this_end)
    prev_matches = run_event_match(event_key, organization_id, prev_start, prev_end)

    this_news = [m for m in this_matches if m["platform_type"] == "news"]
    this_social = [m for m in this_matches if m["platform_type"] != "news"]
    prev_news = [m for m in prev_matches if m["platform_type"] == "news"]
    prev_social = [m for m in prev_matches if m["platform_type"] != "news"]

    this_week_label = _range_label(report_date - timedelta(days=6), report_date)
    prev_week_label = _range_label(prev_report_date - timedelta(days=6), prev_report_date)

    comparison = {
        "yesterday_label": prev_week_label,
        "today_label": this_week_label,
        "news": {
            "yesterday_total": len(prev_news),
            "today_total": len(this_news),
            "yesterday_sentiment": _event_sentiment_counts(prev_news),
            "today_sentiment": _event_sentiment_counts(this_news),
        },
        "social": {
            "yesterday_total": len(prev_social),
            "today_total": len(this_social),
            "yesterday_sentiment": _event_sentiment_counts(prev_social),
            "today_sentiment": _event_sentiment_counts(this_social),
        },
    }

    mobifone_news = [m for m in this_news if m["brand"] == org_name]
    competitor_news: dict[str, list[dict]] = {brand: [] for brand in get_competitor_brands(organization_id)}
    for m in this_news:
        if m["brand"] != org_name:
            competitor_news.setdefault(m["brand"], []).append(m)
    social_mobifone = [m for m in this_social if m["brand"] == org_name]

    brands = [org_name] + get_competitor_brands(organization_id)
    brand_counts = {
        brand: _event_sentiment_counts([m for m in this_matches if m["brand"] == brand]) for brand in brands
    }

    overview_narrative = generate_overview_narrative(org_name, this_news)

    return {
        "org_name": org_name,
        "event_label": event["label"],
        "period_label": this_week_label,
        "comparison": comparison,
        "overview_narrative": overview_narrative,
        "mobifone_news": mobifone_news,
        "competitor_news": competitor_news,
        "social_matches": social_mobifone,
        "brand_counts": brand_counts,
    }


def build_event_weekly_word_report_bytes(user: dict, event_key: str, report_date: date) -> bytes:
    """Weekly variant of build_event_word_report_bytes — 7-day window ending
    on report_date (via weekly_window) vs. the preceding 7-day window, plus
    a brand-vs-brand (MobiFone + each competitor) sentiment summary the
    daily report doesn't have."""
    from platform_app.reporting.event_word_report import build_event_weekly_word_report_bytes as _build_weekly_docx

    data = compute_event_weekly_report_data(user, event_key, report_date)
    return _build_weekly_docx(**data)


@router.get("/report/event/{event_key}/export-word-weekly")
def export_event_report_word_weekly(
    event_key: str,
    report_date: date = Query(default_factory=lambda: datetime.now(timezone.utc).date()),
    user: dict = Depends(get_current_user),
) -> StreamingResponse:
    """Word (.docx) weekly "sự vụ" (topic-event) report — 7-day window
    ending 08:00 (Vietnam time) on `report_date`, plus the preceding week
    for comparison, plus a brand-vs-brand sentiment summary the daily
    report doesn't have."""
    from platform_app.pipeline.event_report import EVENT_DEFINITIONS

    if event_key not in EVENT_DEFINITIONS:
        raise HTTPException(status.HTTP_404_NOT_FOUND, f"Không tìm thấy sự vụ '{event_key}'")

    content = build_event_weekly_word_report_bytes(user, event_key, report_date)
    filename = f"bao-cao-tuan-{event_key}-{report_date.strftime('%Y%m%d')}.docx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/report/event/{event_key}/data-weekly")
def get_event_weekly_report_data(
    event_key: str,
    report_date: date = Query(default_factory=lambda: datetime.now(timezone.utc).date()),
    user: dict = Depends(get_current_user),
) -> dict:
    """JSON preview of the same data export_event_report_word_weekly renders
    to .docx."""
    from platform_app.pipeline.event_report import EVENT_DEFINITIONS

    if event_key not in EVENT_DEFINITIONS:
        raise HTTPException(status.HTTP_404_NOT_FOUND, f"Không tìm thấy sự vụ '{event_key}'")

    return compute_event_weekly_report_data(user, event_key, report_date)


def compute_negative_brand_report_data(user: dict, report_date: date) -> dict:
    """All data build_negative_brand_report_bytes needs, computed once —
    reused by both the Word export and the JSON preview route
    (/report/negative-brand/data-weekly). Unlike the 5G event report, this
    scopes to the org's own-brand documents overall (brand_focus='own'),
    reusing the same daily-social-report infra (_report_scope_between /
    _build_report_from_scope) instead of event_report.py's topic-match
    pipeline. Sections with no backing data anywhere in the system (geo
    hotspots, xử lý/seeding case tracking) render as manual placeholders —
    see build_negative_brand_weekly_word_report_bytes."""
    from platform_app.pipeline.negative_brand_report import (
        generate_negative_overview_narrative,
        summarize_negative_theme,
    )
    from platform_app.reporting.word_report import _pct, _pct_change, weekly_window

    org_name = user["organization_name"]
    prev_report_date = report_date - timedelta(days=7)

    this_start, this_end = weekly_window(report_date)
    prev_start, prev_end = weekly_window(prev_report_date)

    this_scope = _report_scope_between(user, this_start, this_end, None, brand_focus="own")
    prev_scope = _report_scope_between(user, prev_start, prev_end, None, brand_focus="own")
    this_report = _build_report_from_scope(user, this_scope, None) if this_scope else dict(_EMPTY_REPORT)
    prev_report = _build_report_from_scope(user, prev_scope, None) if prev_scope else dict(_EMPTY_REPORT)

    this_channel_split = {"news": 0, "social": 0}
    news_theme = "Không có phản ánh tiêu cực nào trên kênh báo chí online trong tuần này."
    news_pct = "0%"
    news_location: str | None = None
    social_theme = "Không có phản ánh tiêu cực nào trên mạng xã hội trong tuần này."
    social_pct = "0%"
    social_location: str | None = None

    with get_pool().connection() as conn:
        if prev_scope:
            prev_conditions, prev_params, _, _ = prev_scope
            prev_channel_split = _negative_channel_split(conn, prev_conditions, prev_params)
        else:
            prev_channel_split = {"news": 0, "social": 0}

        if this_scope:
            this_conditions, this_params, _, _ = this_scope
            this_channel_split = _negative_channel_split(conn, this_conditions, this_params)

            top_news = _negative_top_topic(conn, this_conditions, this_params, news=True)
            if top_news:
                topic_name, topic_count = top_news
                docs = _negative_topic_docs(conn, this_conditions, this_params, news=True, topic=topic_name)
                news_theme, news_location = summarize_negative_theme(org_name, "kênh báo chí online", docs)
                news_pct = _pct(topic_count, this_channel_split["news"])

            top_social = _negative_top_topic(conn, this_conditions, this_params, news=False)
            if top_social:
                topic_name, topic_count = top_social
                docs = _negative_topic_docs(conn, this_conditions, this_params, news=False, topic=topic_name)
                social_theme, social_location = summarize_negative_theme(org_name, "mạng xã hội", docs)
                social_pct = _pct(topic_count, this_channel_split["social"])

    hotspot_parts = []
    if news_location:
        hotspot_parts.append(f"{news_location} (báo chí)")
    if social_location:
        hotspot_parts.append(f"{social_location} (mạng xã hội)")
    hotspot_text = ", ".join(hotspot_parts) if hotspot_parts else "—"

    this_week_label = _range_label(report_date - timedelta(days=6), report_date)
    prev_week_label = _range_label(prev_report_date - timedelta(days=6), prev_report_date)

    summary_rows = [
        {
            "stt": "1.1",
            "label": "Tổng thông tin thu thập và kiểm soát",
            "prev": prev_report["total_posts"],
            "this": this_report["total_posts"],
            "pct": None,
            "compare": _pct_change(prev_report["total_posts"], this_report["total_posts"]),
            "bold": True,
        },
        {
            "stt": "1.2",
            "label": "Tổng thông tin tiêu cực (đã kiểm soát và cảnh báo)",
            "prev": prev_report["sentiment_negative"],
            "this": this_report["sentiment_negative"],
            "pct": _pct(this_report["sentiment_negative"], this_report["total_posts"]),
            "compare": _pct_change(prev_report["sentiment_negative"], this_report["sentiment_negative"]),
            "bold": True,
        },
        {
            "stt": "-",
            "label": f"Tin tiêu cực về {org_name} trên kênh Báo chí online",
            "prev": prev_channel_split["news"],
            "this": this_channel_split["news"],
            "pct": _pct(this_channel_split["news"], this_report["sentiment_negative"]),
            "compare": _pct_change(prev_channel_split["news"], this_channel_split["news"]),
        },
        {
            "stt": "-",
            "label": f"Tin tiêu cực về {org_name} trên mạng xã hội",
            "prev": prev_channel_split["social"],
            "this": this_channel_split["social"],
            "pct": _pct(this_channel_split["social"], this_report["sentiment_negative"]),
            "compare": _pct_change(prev_channel_split["social"], this_channel_split["social"]),
        },
        {
            "stt": "-",
            "label": "Ghi nhận và phối hợp xử lý (trường hợp)",
            "prev": None,
            "this": None,
            "pct": None,
            "compare": None,
        },
        {
            "stt": "1.3",
            "label": "Tổng thông tin tích cực",
            "prev": prev_report["sentiment_positive"],
            "this": this_report["sentiment_positive"],
            "pct": None,
            "compare": _pct_change(prev_report["sentiment_positive"], this_report["sentiment_positive"]),
            "bold": True,
        },
        {
            "stt": "1.4",
            "label": "Tổng thông tin trung tính",
            "prev": prev_report["sentiment_neutral"],
            "this": this_report["sentiment_neutral"],
            "pct": None,
            "compare": _pct_change(prev_report["sentiment_neutral"], this_report["sentiment_neutral"]),
            "bold": True,
        },
    ]

    comparison = {
        "total_prev": prev_report["total_posts"],
        "total_this": this_report["total_posts"],
        "negative_prev": prev_report["sentiment_negative"],
        "negative_this": this_report["sentiment_negative"],
        "negative_news_this": this_channel_split["news"],
        "negative_social_this": this_channel_split["social"],
    }
    overview_narrative = generate_negative_overview_narrative(org_name, comparison, news_theme, social_theme)

    return {
        "org_name": org_name,
        "period_label": this_week_label,
        "period_prev_label": prev_week_label,
        "summary_rows": summary_rows,
        "news_theme": news_theme,
        "news_pct": news_pct,
        "social_theme": social_theme,
        "social_pct": social_pct,
        "hotspot_text": hotspot_text,
        "overview_narrative": overview_narrative,
    }


def build_negative_brand_report_bytes(user: dict, report_date: date) -> bytes:
    """Weekly "negative brand mentions" report."""
    from platform_app.reporting.negative_report import build_negative_brand_weekly_word_report_bytes as _build_docx

    data = compute_negative_brand_report_data(user, report_date)
    return _build_docx(**data)


@router.get("/report/negative-brand/export-word-weekly")
def export_negative_brand_report_word_weekly(
    report_date: date = Query(default_factory=lambda: datetime.now(timezone.utc).date()),
    user: dict = Depends(get_current_user),
) -> StreamingResponse:
    """Word (.docx) weekly "negative brand mentions" report — org-wide
    own-brand negative content, 7-day window ending 08:00 (Vietnam time) on
    `report_date`, vs. the preceding week."""
    content = build_negative_brand_report_bytes(user, report_date)
    filename = f"bao-cao-tuan-tieu-cuc-{report_date.strftime('%Y%m%d')}.docx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/report/negative-brand/data-weekly")
def get_negative_brand_report_data(
    report_date: date = Query(default_factory=lambda: datetime.now(timezone.utc).date()),
    user: dict = Depends(get_current_user),
) -> dict:
    """JSON preview of the same data export_negative_brand_report_word_weekly
    renders to .docx."""
    return compute_negative_brand_report_data(user, report_date)


def build_monthly_brand_report_bytes(user: dict, report_date: date) -> bytes:
    """Monthly "Báo cáo tháng thương hiệu MobiFone" — a balanced overview
    (positive + negative + neutral), NOT negative-only (that's
    build_negative_brand_report_bytes, a separate weekly report). Reuses
    build_word_report_bytes_between unchanged — it's already generic over
    any [period_start, period_end) window, exactly the same template the
    daily report and report-email job already use (tổng quan + sentiment
    pie + topic breakdown + top negative/positive posts)."""
    from platform_app.reporting.word_report import monthly_window

    org_name = user["organization_name"]
    period_start, period_end = monthly_window(report_date)
    period_label = f"THÁNG {_range_label(report_date - timedelta(days=29), report_date)}"
    return build_word_report_bytes_between(
        user, org_name, report_date, period_start, period_end, None, period_label=period_label
    )


@router.get("/report/export-word-monthly")
def export_monthly_brand_report_word(
    report_date: date = Query(default_factory=lambda: datetime.now(timezone.utc).date()),
    user: dict = Depends(get_current_user),
) -> StreamingResponse:
    """Word (.docx) monthly brand report — same balanced template as the
    daily report (/report/export-word), 30-day window ending 08:00
    (Vietnam time) on `report_date` instead of 1 day."""
    content = build_monthly_brand_report_bytes(user, report_date)
    filename = f"bao-cao-thang-{report_date.strftime('%Y%m%d')}.docx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def compute_competitor_channel_report_data(user: dict, report_date: date) -> dict:
    """Weekly "đối thủ cùng ngành" report data."""
    from platform_app.reporting.word_report import weekly_window

    return _compute_competitor_channel_report_data(
        user, report_date, window_fn=weekly_window, period_days=7, period_unit_label="TUẦN"
    )


def compute_competitor_channel_report_data_monthly(user: dict, report_date: date) -> dict:
    """Monthly variant of compute_competitor_channel_report_data — identical
    structure, 30-day window instead of 7-day (monthly_window). Requested
    as "Báo cáo đối thủ ++ tháng", the pre-existing catalog stub name for
    exactly this weekly report on a monthly cadence."""
    from platform_app.reporting.word_report import monthly_window

    return _compute_competitor_channel_report_data(
        user, report_date, window_fn=monthly_window, period_days=30, period_unit_label="THÁNG"
    )


def build_competitor_channel_report_bytes(user: dict, report_date: date) -> bytes:
    """Weekly "đối thủ cùng ngành" report — Word export."""
    from platform_app.reporting.competitor_word_report import build_competitor_weekly_word_report_bytes as _build_docx

    data = compute_competitor_channel_report_data(user, report_date)
    return _build_docx(**data)


def build_competitor_channel_monthly_report_bytes(user: dict, report_date: date) -> bytes:
    """Monthly variant of build_competitor_channel_report_bytes — Word export."""
    from platform_app.reporting.competitor_word_report import build_competitor_weekly_word_report_bytes as _build_docx

    data = compute_competitor_channel_report_data_monthly(user, report_date)
    return _build_docx(**data)


def _compute_competitor_channel_report_data(
    user: dict, report_date: date, *, window_fn, period_days: int, period_unit_label: str
) -> dict:
    """Shared impl for the weekly and monthly "đối thủ cùng ngành" report —
    MobiFone + its tracked competitors treated as peer brands (unlike
    build_negative_brand_report_bytes, which only looks at MobiFone's own
    content). Sentiment breakdown reuses documents.classification_sentiment
    (no new LLM call, unlike the 5G event report) — only the 2 bullet
    summaries in mục 1 need an LLM."""
    from platform_app.pipeline.competitor_report import summarize_brand_bullets
    from platform_app.pipeline.event_report import get_competitor_brands
    from platform_app.reporting.word_report import _fmt

    org_name = user["organization_name"]
    organization_id = user["organization_id"]
    brands = [org_name, *get_competitor_brands(organization_id)]

    this_start, this_end = window_fn(report_date)
    this_period_label = f"{period_unit_label} {_range_label(report_date - timedelta(days=period_days - 1), report_date)}"

    brand_counts = {b: {"positive": 0, "neutral": 0, "negative": 0} for b in brands}
    channel_breakdowns = {b: {"Facebook": 0, "News": 0, "Forum": 0} for b in brands}
    competitor_posts: dict[str, dict[str, list[dict]]] = {b: {"positive": [], "negative": []} for b in brands}
    docs_by_brand_positive: dict[str, list[dict]] = {}
    docs_by_brand_negative: dict[str, list[dict]] = {}

    brand_scope = _brand_scope_condition(user, this_start, this_end)
    if brand_scope:
        conditions, params = brand_scope
        with get_pool().connection() as conn:
            brand_counts = _brand_sentiment_counts(conn, conditions, params, brands)
            for brand in brands:
                channel_breakdowns[brand] = _brand_channel_breakdown(conn, conditions, params, brand)
                pos_posts = _brand_top_posts(conn, conditions, params, brand=brand, sentiment="positive")
                neg_posts = _brand_top_posts(conn, conditions, params, brand=brand, sentiment="negative")
                docs_by_brand_positive[brand] = pos_posts
                docs_by_brand_negative[brand] = neg_posts
                competitor_posts[brand] = {"positive": pos_posts, "negative": neg_posts}

    own_scope = _report_scope_between(user, this_start, this_end, None, brand_focus="own")
    own_report = _build_report_from_scope(user, own_scope, None) if own_scope else dict(_EMPTY_REPORT)

    positive_bullets = summarize_brand_bullets(brands, "tích cực", docs_by_brand_positive)
    negative_bullets = summarize_brand_bullets(brands, "tiêu cực", docs_by_brand_negative)

    # mục 4's 2 bullet tính trực tiếp bằng Python (argmax trên số liệu thật) —
    # đáng tin cậy hơn để LLM suy luận từ 1 con số duy nhất.
    channel_bullets: list[str] = []
    fb_leader = max(brands, key=lambda b: channel_breakdowns[b]["Facebook"], default=None)
    if fb_leader and channel_breakdowns[fb_leader]["Facebook"] > 0:
        max_fb = channel_breakdowns[fb_leader]["Facebook"]
        channel_bullets.append(f"-  Xu hướng khách hàng nhắc đến {fb_leader.upper()} nhiều nhất trên kênh Facebook.")
        channel_bullets.append(
            f"-  Số bài đăng trên kênh Facebook của {fb_leader.upper()} cao nhất trong tất cả các nhà mạng "
            f"với {_fmt(max_fb)} bài đăng."
        )

    return {
        "org_name": org_name,
        "period_label": this_period_label,
        "brands": brands,
        "brand_counts": brand_counts,
        "positive_bullets": positive_bullets,
        "negative_bullets": negative_bullets,
        "own_positive_posts": own_report["positive_posts"],
        "own_negative_posts": own_report["negative_posts"],
        "competitor_posts": competitor_posts,
        "channel_breakdowns": channel_breakdowns,
        "channel_bullets": channel_bullets,
    }


@router.get("/report/competitor-channels/export-word-weekly")
def export_competitor_channel_report_word_weekly(
    report_date: date = Query(default_factory=lambda: datetime.now(timezone.utc).date()),
    user: dict = Depends(get_current_user),
) -> StreamingResponse:
    """Word (.docx) weekly "đối thủ cùng ngành" report — MobiFone + tracked
    competitors as peer brands, 7-day window ending 08:00 (Vietnam time) on
    `report_date`."""
    content = build_competitor_channel_report_bytes(user, report_date)
    filename = f"bao-cao-tuan-doi-thu-{report_date.strftime('%Y%m%d')}.docx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/report/competitor-channels/export-word-monthly")
def export_competitor_channel_report_word_monthly(
    report_date: date = Query(default_factory=lambda: datetime.now(timezone.utc).date()),
    user: dict = Depends(get_current_user),
) -> StreamingResponse:
    """Word (.docx) monthly "đối thủ cùng ngành" report — same report as the
    weekly export, 30-day window instead of 7-day."""
    content = build_competitor_channel_monthly_report_bytes(user, report_date)
    filename = f"bao-cao-thang-doi-thu-{report_date.strftime('%Y%m%d')}.docx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/report/competitor-channels/data-weekly")
def get_competitor_channel_report_data(
    report_date: date = Query(default_factory=lambda: datetime.now(timezone.utc).date()),
    user: dict = Depends(get_current_user),
) -> dict:
    """JSON preview of the same data export_competitor_channel_report_word_weekly
    renders to .docx."""
    return compute_competitor_channel_report_data(user, report_date)


@router.get("/report/competitor-channels/data-monthly")
def get_competitor_channel_report_data_monthly(
    report_date: date = Query(default_factory=lambda: datetime.now(timezone.utc).date()),
    user: dict = Depends(get_current_user),
) -> dict:
    """JSON preview of the same data export_competitor_channel_report_word_monthly
    renders to .docx."""
    return compute_competitor_channel_report_data_monthly(user, report_date)


# ---------------------------------------------------------------------------
# Settings — classify_mode is per-organization (pipeline_settings.organization_id);
# every member of the org can see the current mode, but only org_main (the
# owner account) can change it, since it directly controls LLM spend.
# ---------------------------------------------------------------------------


@router.get("/settings/classify-mode")
def get_org_classify_mode(user: dict = Depends(get_current_user)) -> dict:
    return {"mode": get_classify_mode(user["organization_id"]), "modes": list(VALID_MODES)}


@router.patch("/settings/classify-mode")
def update_org_classify_mode(body: dict, user: dict = Depends(get_current_user)) -> dict:
    if user["role"] != "org_main":
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Chỉ Tài khoản Chủ mới có thể đổi cài đặt này")
    mode = body.get("mode")
    if mode not in VALID_MODES:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"classify_mode không hợp lệ: {mode}")
    set_classify_mode(mode, user["organization_id"])
    return {"mode": mode, "modes": list(VALID_MODES)}


_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


@router.get("/settings/report-email")
def get_report_email_setting(user: dict = Depends(get_current_user)) -> dict:
    with get_pool().connection() as conn:
        row = conn.execute(
            "SELECT recipient_email, cc_emails, enabled FROM organization_report_email WHERE organization_id = %s",
            (user["organization_id"],),
        ).fetchone()
    if row is None:
        return {"recipient_email": None, "cc_emails": [], "enabled": False}
    return row


@router.patch("/settings/report-email")
def update_report_email_setting(body: dict, user: dict = Depends(get_current_user)) -> dict:
    if user["role"] != "org_main":
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Chỉ Tài khoản Chủ mới có thể đổi cài đặt này")

    recipient_email = (body.get("recipient_email") or "").strip()
    cc_emails = [e.strip() for e in (body.get("cc_emails") or []) if e.strip()]
    enabled = bool(body.get("enabled", True))

    if enabled or recipient_email:
        if not _EMAIL_RE.match(recipient_email):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Email người nhận không hợp lệ: {recipient_email}")
        for cc in cc_emails:
            if not _EMAIL_RE.match(cc):
                raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Email CC không hợp lệ: {cc}")

    with get_pool().connection() as conn:
        row = conn.execute(
            """
            INSERT INTO organization_report_email (organization_id, recipient_email, cc_emails, enabled)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (organization_id) DO UPDATE SET
                recipient_email = EXCLUDED.recipient_email,
                cc_emails = EXCLUDED.cc_emails,
                enabled = EXCLUDED.enabled,
                updated_at = now()
            RETURNING recipient_email, cc_emails, enabled
            """,
            (user["organization_id"], recipient_email, cc_emails, enabled),
        ).fetchone()
    return row


@router.post("/report/send-email")
def send_report_email_now(
    days: int = Query(default=7, ge=1, le=365),
    entity: str | None = Query(default=None),
    user: dict = Depends(get_current_user),
) -> dict:
    """Manual "send now" — uses whatever recipient/cc is already saved in
    Settings, regardless of the `enabled` toggle (that toggle only gates the
    automated daily job, not an explicit manual trigger)."""
    from platform_app.notifications.email import EmailNotConfigured, send_email_with_attachment

    with get_pool().connection() as conn:
        row = conn.execute(
            "SELECT recipient_email, cc_emails FROM organization_report_email WHERE organization_id = %s",
            (user["organization_id"],),
        ).fetchone()
    if row is None or not row["recipient_email"]:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Chưa cấu hình email nhận báo cáo trong Cài đặt")

    period_end = datetime.now(timezone.utc)
    period_start = period_end - timedelta(days=days)
    content = build_word_report_bytes_between(
        user, user["organization_name"], period_end.date(), period_start, period_end, entity
    )
    date_label = period_end.strftime("%d/%m/%Y")
    try:
        send_email_with_attachment(
            to=row["recipient_email"],
            cc=row["cc_emails"],
            subject=f"[{user['organization_name']}] Báo cáo mạng xã hội ngày {date_label}",
            body_text=(
                f"Chào {user['organization_name']},\n\n"
                f"Đính kèm là báo cáo tổng hợp mạng xã hội ({days} ngày gần nhất, gửi thủ công).\n\n"
                "Email này được gửi tự động, vui lòng không trả lời."
            ),
            attachment_bytes=content,
            attachment_filename=f"bao-cao-{period_end.strftime('%Y%m%d-%H%M')}.docx",
        )
    except EmailNotConfigured as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "SMTP chưa được cấu hình trên hệ thống") from exc
    except Exception as exc:
        raise HTTPException(status.HTTP_502_BAD_GATEWAY, f"Gửi email thất bại: {exc}") from exc

    return {"sent_to": row["recipient_email"], "cc": row["cc_emails"]}


# ---------------------------------------------------------------------------
# Documents — crawled posts from the caller's own crawl_targets, optionally
# narrowed to their tracked entities/keywords (organization_entities /
# organization_keywords are a UI preference only — keyword_filter.py and
# entity_match.py tag every document against the full global catalog, so
# "entity"/"keyword" here just filter to documents that happen to match a
# given canonical_name/matched_keyword, same as the internal /report page).
# ---------------------------------------------------------------------------


def _document_list_conditions(user: dict) -> tuple[list[str], list]:
    conditions = ["ct.organization_id = %s"]
    params: list = [user["organization_id"]]
    if user["role"] == "org_sub":
        ids = user["accessible_target_ids"] or []
        conditions.append("d.target_id = ANY(%s)")
        params.append(ids)
    return conditions, params


def _apply_keyword_entity_filters(
    conditions: list[str],
    params: list,
    search_text: str | None,
    entity: str | None,
    entity_exact: bool,
) -> None:
    """Shared free-text narrowing for the plain document list AND every
    accordion aggregate (counts/sentiment-counts/growth/network) so a
    collapsed section's count always matches what expanding it will show.
    `entity_exact=False` substring-matches canonical_name (mirrors
    opencrawler's Accordion_View "Entity" box); `True` requires an exact
    (case-insensitive) match — used by the dropdown-driven /documents list."""
    if search_text:
        conditions.append("(d.topic ILIKE %s OR d.content ILIKE %s)")
        pattern = f"%{search_text}%"
        params.extend([pattern, pattern])
    if entity:
        conditions.append(
            "EXISTS (SELECT 1 FROM document_entities de WHERE de.document_id = d.id AND de.canonical_name ILIKE %s)"
        )
        params.append(entity if entity_exact else f"%{entity}%")


def _apply_days_filter(
    conditions: list[str],
    params: list,
    days: int | None,
    date_from: str | None = None,
    date_to: str | None = None,
) -> None:
    """Restricts to documents published in a time window. An explicit
    `date_from`/`date_to` (custom range, "YYYY-MM-DD") takes priority over
    the `days` preset when both are supplied; with neither, no time filter
    is applied (all-time) — matches the accordion/document list's default of
    showing everything until the user opts into a window."""
    if date_from or date_to:
        if date_from:
            conditions.append("d.published_at >= %s")
            params.append(datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=timezone.utc))
        if date_to:
            conditions.append("d.published_at < %s")
            params.append(
                datetime.strptime(date_to, "%Y-%m-%d").replace(tzinfo=timezone.utc) + timedelta(days=1)
            )
    elif days:
        conditions.append("d.published_at >= %s")
        params.append(datetime.now(timezone.utc) - timedelta(days=days))


@router.get("/documents", response_model=DocumentListResponse)
def list_documents(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    entity: str | None = Query(default=None),
    entity_exact: bool = Query(default=True),
    keyword: str | None = Query(default=None),
    platform_type: str | None = Query(default=None),
    sentiment: str | None = Query(default=None),
    search: str | None = Query(default=None),
    days: int | None = Query(default=None, ge=1, le=3650),
    date_from: str | None = Query(default=None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    date_to: str | None = Query(default=None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    user: dict = Depends(get_current_user),
) -> dict:
    conditions, params = _document_list_conditions(user)
    if user["role"] == "org_sub" and not (user["accessible_target_ids"] or []):
        return {"items": [], "total": 0}

    if platform_type:
        conditions.append("d.platform_type = %s")
        params.append(platform_type)
    if sentiment == "competitor":
        conditions.append("d.brand_focus = 'competitor'")
    elif sentiment == "unclassified":
        conditions.append(f"d.classification_sentiment IS NULL AND {_NOT_COMPETITOR}")
    elif sentiment:
        conditions.append("d.classification_sentiment = %s AND " + _NOT_COMPETITOR)
        params.append(sentiment)
    if search:
        conditions.append("(d.topic ILIKE %s OR d.content ILIKE %s)")
        pattern = f"%{search}%"
        params.extend([pattern, pattern])
    _apply_keyword_entity_filters(conditions, params, None, entity, entity_exact)
    _apply_days_filter(conditions, params, days, date_from, date_to)
    if keyword:
        conditions.append("d.matched_keywords ? %s")
        params.append(keyword)

    where_clause = " AND ".join(conditions)

    with get_pool().connection() as conn:
        total = conn.execute(
            f"""
            SELECT COUNT(*) AS n
            FROM documents d
            JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE {where_clause}
            """,
            params,
        ).fetchone()["n"]

        rows = conn.execute(
            f"""
            SELECT d.id, d.platform_type, d.source_type, ct.display_name AS target_name,
                   d.author, d.topic,
                   LEFT(d.content, 220) AS content_snippet,
                   d.url, d.published_at,
                   d.like_count, d.comment_count, d.reaction_count, d.share_count,
                   d.keyword_status, d.matched_keywords,
                   d.classification_category, d.classification_sentiment, d.classification_severity,
                   COALESCE(
                       (SELECT array_agg(DISTINCT de.canonical_name)
                        FROM document_entities de WHERE de.document_id = d.id AND de.concept_id != '__none__'
                          AND {_tracked_entity_condition("de")}),
                       ARRAY[]::text[]
                   ) AS entities
            FROM documents d
            JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE {where_clause}
            ORDER BY d.published_at DESC NULLS LAST, d.id DESC
            LIMIT %s OFFSET %s
            """,
            [user["organization_id"], *params, page_size, (page - 1) * page_size],
        ).fetchall()

    return {"items": rows, "total": total}


@router.get("/documents/{document_id}", response_model=DocumentDetailOut)
def get_document(document_id: int, user: dict = Depends(get_current_user)) -> dict:
    with get_pool().connection() as conn:
        row = conn.execute(
            f"""
            SELECT d.id, d.target_id, d.platform_type, d.source_type, ct.display_name AS target_name,
                   ct.organization_id,
                   d.author, d.topic, d.content, d.url, d.published_at,
                   d.images, d.videos,
                   d.like_count, d.comment_count, d.reaction_count, d.share_count, d.reactions,
                   d.keyword_status, d.matched_keywords,
                   d.classification_category, d.classification_sentiment, d.classification_sentiment_source,
                   d.classification_severity, d.classification_reasoning,
                   d.classification_text_summary, d.classification_image_summary,
                   COALESCE(
                       (SELECT array_agg(DISTINCT de.canonical_name)
                        FROM document_entities de WHERE de.document_id = d.id AND de.concept_id != '__none__'
                          AND {_tracked_entity_condition("de")}),
                       ARRAY[]::text[]
                   ) AS entities
            FROM documents d
            JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE d.id = %s
            """,
            (user["organization_id"], document_id),
        ).fetchone()

    if row is None or row["organization_id"] != user["organization_id"]:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy bài viết")
    if user["role"] == "org_sub" and row["target_id"] not in (user["accessible_target_ids"] or []):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy bài viết")

    return row


@router.get("/documents/{document_id}/comments", response_model=list[DocumentCommentOut])
def get_document_comments(document_id: int, user: dict = Depends(get_current_user)) -> list[dict]:
    with get_pool().connection() as conn:
        doc = conn.execute(
            """
            SELECT d.target_id, ct.organization_id
            FROM documents d JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE d.id = %s
            """,
            (document_id,),
        ).fetchone()
        if doc is None or doc["organization_id"] != user["organization_id"]:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy bài viết")
        if user["role"] == "org_sub" and doc["target_id"] not in (user["accessible_target_ids"] or []):
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy bài viết")

        return conn.execute(
            """
            SELECT author, text, created_at, depth
            FROM document_comments
            WHERE document_id = %s
            ORDER BY created_at NULLS LAST, id
            """,
            (document_id,),
        ).fetchall()


# ---------------------------------------------------------------------------
# Accordion view — alternate browse UI mirroring opencrawler's Accordion_View:
# one collapsible section per platform_type, sentiment sub-tabs, a growth
# chart + entity co-occurrence network per section, and (per selected post)
# related posts sharing an entity. All aggregates share the same
# org/keyword/entity narrowing as /documents above so a collapsed section's
# count always matches what expanding it will show.
# ---------------------------------------------------------------------------


@router.get("/documents/accordion/counts", response_model=AccordionCategoryCounts)
def accordion_category_counts(
    search: str | None = Query(default=None),
    entity: str | None = Query(default=None),
    entity_exact: bool = Query(default=False),
    days: int | None = Query(default=None, ge=1, le=3650),
    date_from: str | None = Query(default=None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    date_to: str | None = Query(default=None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    user: dict = Depends(get_current_user),
) -> dict:
    conditions, params = _document_list_conditions(user)
    if user["role"] == "org_sub" and not (user["accessible_target_ids"] or []):
        return {p: 0 for p in VALID_PLATFORM_TYPES}
    _apply_keyword_entity_filters(conditions, params, search, entity, entity_exact)
    _apply_days_filter(conditions, params, days, date_from, date_to)
    where_clause = " AND ".join(conditions)

    with get_pool().connection() as conn:
        rows = conn.execute(
            f"""
            SELECT d.platform_type, COUNT(*) AS n
            FROM documents d
            JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE {where_clause}
            GROUP BY d.platform_type
            """,
            params,
        ).fetchall()

    counts = {p: 0 for p in VALID_PLATFORM_TYPES}
    for row in rows:
        if row["platform_type"] in counts:
            counts[row["platform_type"]] = row["n"]
    return counts


_NOT_COMPETITOR = "(d.brand_focus IS DISTINCT FROM 'competitor')"
_SENTIMENT_BUCKET_SQL = {
    "positive": f"d.classification_sentiment = 'positive' AND {_NOT_COMPETITOR}",
    "negative": f"d.classification_sentiment = 'negative' AND {_NOT_COMPETITOR}",
    "neutral": f"d.classification_sentiment = 'neutral' AND {_NOT_COMPETITOR}",
    "unclassified": f"d.classification_sentiment IS NULL AND {_NOT_COMPETITOR}",
    "competitor": "d.brand_focus = 'competitor'",
}


@router.get("/documents/accordion/sentiment-counts", response_model=AccordionSentimentCounts)
def accordion_sentiment_counts(
    platform_type: str = Query(...),
    search: str | None = Query(default=None),
    entity: str | None = Query(default=None),
    entity_exact: bool = Query(default=False),
    days: int | None = Query(default=None, ge=1, le=3650),
    date_from: str | None = Query(default=None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    date_to: str | None = Query(default=None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    user: dict = Depends(get_current_user),
) -> dict:
    if platform_type not in VALID_PLATFORM_TYPES:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "platform_type không hợp lệ")

    conditions, params = _document_list_conditions(user)
    if user["role"] == "org_sub" and not (user["accessible_target_ids"] or []):
        return {k: 0 for k in _SENTIMENT_BUCKET_SQL}
    conditions.append("d.platform_type = %s")
    params.append(platform_type)
    _apply_keyword_entity_filters(conditions, params, search, entity, entity_exact)
    _apply_days_filter(conditions, params, days, date_from, date_to)
    where_clause = " AND ".join(conditions)

    with get_pool().connection() as conn:
        counts: dict[str, int] = {}
        for key, cond in _SENTIMENT_BUCKET_SQL.items():
            counts[key] = conn.execute(
                f"""
                SELECT COUNT(*) AS n
                FROM documents d
                JOIN crawl_targets ct ON ct.id = d.target_id
                WHERE {where_clause} AND {cond}
                """,
                params,
            ).fetchone()["n"]
    return counts


@router.get("/documents/accordion/growth", response_model=list[EngagementGrowthPoint])
def accordion_engagement_growth(
    platform_type: str = Query(...),
    search: str | None = Query(default=None),
    entity: str | None = Query(default=None),
    entity_exact: bool = Query(default=False),
    days: int | None = Query(default=None, ge=1, le=3650),
    date_from: str | None = Query(default=None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    date_to: str | None = Query(default=None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    user: dict = Depends(get_current_user),
) -> list[dict]:
    """Combined engagement-growth curve (summed across every document
    currently matching the filters) — buckets by the actual crawl timestamp
    (truncated to the hour), so the x-axis is a real calendar time, not an
    abstract offset. Sparse/empty until repeat crawls accumulate snapshot
    rows (see document_engagement_snapshots + migration 0010)."""
    if platform_type not in VALID_PLATFORM_TYPES:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "platform_type không hợp lệ")

    conditions, params = _document_list_conditions(user)
    if user["role"] == "org_sub" and not (user["accessible_target_ids"] or []):
        return []
    conditions.append("d.platform_type = %s")
    params.append(platform_type)
    _apply_keyword_entity_filters(conditions, params, search, entity, entity_exact)
    _apply_days_filter(conditions, params, days, date_from, date_to)
    where_clause = " AND ".join(conditions)

    with get_pool().connection() as conn:
        return conn.execute(
            f"""
            SELECT date_trunc('hour', s.crawled_at) AS bucket,
                   SUM(s.like_count) AS like_count,
                   SUM(s.comment_count) AS comment_count,
                   SUM(s.reaction_count) AS reaction_count,
                   SUM(s.share_count) AS share_count
            FROM document_engagement_snapshots s
            JOIN documents d ON d.id = s.document_id
            JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE {where_clause}
            GROUP BY bucket
            ORDER BY bucket
            """,
            params,
        ).fetchall()


@router.get("/documents/accordion/network", response_model=EntityNetworkResponse)
def accordion_entity_network(
    platform_type: str = Query(...),
    search: str | None = Query(default=None),
    entity: str | None = Query(default=None),
    entity_exact: bool = Query(default=False),
    days: int | None = Query(default=None, ge=1, le=3650),
    date_from: str | None = Query(default=None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    date_to: str | None = Query(default=None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    max_nodes: int = Query(default=20, ge=2, le=50),
    user: dict = Depends(get_current_user),
) -> dict:
    """Entity co-occurrence network for every document currently matching the
    accordion filters in one platform_type — nodes are the top `max_nodes`
    canonical entities by how many matched documents mention them; edges are
    how many matched documents mention BOTH ends of the pair."""
    if platform_type not in VALID_PLATFORM_TYPES:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "platform_type không hợp lệ")

    conditions, params = _document_list_conditions(user)
    if user["role"] == "org_sub" and not (user["accessible_target_ids"] or []):
        return {"nodes": [], "edges": []}
    conditions.append("d.platform_type = %s")
    params.append(platform_type)
    _apply_keyword_entity_filters(conditions, params, search, entity, entity_exact)
    _apply_days_filter(conditions, params, days, date_from, date_to)
    where_clause = " AND ".join(conditions)

    with get_pool().connection() as conn:
        nodes = conn.execute(
            f"""
            SELECT de.canonical_name, COUNT(DISTINCT de.document_id) AS post_count
            FROM document_entities de
            JOIN documents d ON d.id = de.document_id
            JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE de.concept_id != '__none__' AND {where_clause} AND {_tracked_entity_condition("de")}
            GROUP BY de.canonical_name
            ORDER BY post_count DESC
            LIMIT %s
            """,
            [*params, user["organization_id"], max_nodes],
        ).fetchall()

        names = [n["canonical_name"] for n in nodes]
        edges = []
        if len(names) >= 2:
            edges = conn.execute(
                f"""
                SELECT de1.canonical_name AS source, de2.canonical_name AS target,
                       COUNT(DISTINCT de1.document_id) AS weight
                FROM document_entities de1
                JOIN document_entities de2
                    ON de2.document_id = de1.document_id AND de2.canonical_name > de1.canonical_name
                JOIN documents d ON d.id = de1.document_id
                JOIN crawl_targets ct ON ct.id = d.target_id
                WHERE de1.canonical_name = ANY(%s) AND de2.canonical_name = ANY(%s) AND {where_clause}
                GROUP BY de1.canonical_name, de2.canonical_name
                """,
                [names, names, *params],
            ).fetchall()

    return {"nodes": nodes, "edges": edges}


@router.get("/documents/{document_id}/related", response_model=list[RelatedDocumentItem])
def get_related_documents(
    document_id: int,
    sentiment: list[str] | None = Query(default=None),
    limit: int = Query(default=10, ge=1, le=50),
    user: dict = Depends(get_current_user),
) -> list[dict]:
    """Other documents sharing at least one canonical entity with
    `document_id`, ranked by shared-entity count then recency — powers the
    accordion detail panel's "bài viết liên quan" block."""
    with get_pool().connection() as conn:
        doc = conn.execute(
            """
            SELECT d.target_id, ct.organization_id
            FROM documents d JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE d.id = %s
            """,
            (document_id,),
        ).fetchone()
        if doc is None or doc["organization_id"] != user["organization_id"]:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy bài viết")
        if user["role"] == "org_sub" and doc["target_id"] not in (user["accessible_target_ids"] or []):
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy bài viết")

        conditions, params = _document_list_conditions(user)
        conditions.append("de.document_id != %s")
        params.append(document_id)
        if sentiment:
            conditions.append("d.classification_sentiment = ANY(%s)")
            params.append(sentiment)
        where_clause = " AND ".join(conditions)

        return conn.execute(
            f"""
            SELECT d.id, d.platform_type, ct.display_name AS target_name, d.topic,
                   LEFT(d.content, 220) AS content_snippet, d.published_at,
                   d.classification_sentiment, COUNT(*) AS shared_entities
            FROM document_entities de
            JOIN document_entities mine
                ON mine.canonical_name = de.canonical_name AND mine.document_id = %s
            JOIN documents d ON d.id = de.document_id
            JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE de.concept_id != '__none__' AND {where_clause} AND {_tracked_entity_condition("de")}
            GROUP BY d.id, d.platform_type, ct.display_name, d.topic, d.content, d.published_at,
                     d.classification_sentiment
            ORDER BY shared_entities DESC, d.published_at DESC NULLS LAST
            LIMIT %s
            """,
            [document_id, *params, user["organization_id"], limit],
        ).fetchall()


@router.get("/documents/{document_id}/entity-network", response_model=EntityNetworkResponse)
def get_document_entity_network(
    document_id: int,
    focus: str | None = Query(default=None),
    focus_exact: bool = Query(default=False),
    user: dict = Depends(get_current_user),
) -> dict:
    """Co-occurrence network for THIS document's own entities — nodes are the
    entities tagged on this one document; an edge's weight is how many OTHER
    documents (within the caller's organization) tag that same pair. Distinct
    from /documents/accordion/network (which aggregates over every document
    matching a category/filter, not just one document's own entities) —
    mirrors opencrawler's post_detail.render_entity_relationship_graph.
    Renders nothing on the frontend for fewer than 2 entities.

    `focus` mirrors the page-level Entity search box: if it resolves to one
    of THIS document's own entities (same canonical_name matching as
    /documents' entity filter), the graph narrows to a star centered on that
    entity — only ITS relationships are returned, not every other pair —
    since once an analyst searched for a specific entity, only its
    relationships answer the question."""
    with get_pool().connection() as conn:
        doc = conn.execute(
            """
            SELECT d.target_id, ct.organization_id
            FROM documents d JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE d.id = %s
            """,
            (document_id,),
        ).fetchone()
        if doc is None or doc["organization_id"] != user["organization_id"]:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy bài viết")
        if user["role"] == "org_sub" and doc["target_id"] not in (user["accessible_target_ids"] or []):
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy bài viết")

        own_names = [
            row["canonical_name"]
            for row in conn.execute(
                f"""
                SELECT DISTINCT canonical_name FROM document_entities
                WHERE document_id = %s AND concept_id != '__none__' AND {_tracked_entity_condition("document_entities")}
                """,
                (document_id, user["organization_id"]),
            ).fetchall()
        ]
        if len(own_names) < 2:
            return {"nodes": [{"canonical_name": n, "post_count": 1} for n in own_names], "edges": []}

        resolved_focus = None
        if focus and focus.strip():
            q = focus.strip().lower()
            for name in own_names:
                if (focus_exact and name.lower() == q) or (not focus_exact and q in name.lower()):
                    resolved_focus = name
                    break

        conditions, params = _document_list_conditions(user)
        where_clause = " AND ".join(conditions)

        nodes = conn.execute(
            f"""
            SELECT de.canonical_name, COUNT(DISTINCT de.document_id) AS post_count
            FROM document_entities de
            JOIN documents d ON d.id = de.document_id
            JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE de.canonical_name = ANY(%s) AND {where_clause}
            GROUP BY de.canonical_name
            """,
            [own_names, *params],
        ).fetchall()

        if resolved_focus:
            others = [n for n in own_names if n != resolved_focus]
            edges = conn.execute(
                f"""
                SELECT %s AS source, de2.canonical_name AS target,
                       COUNT(DISTINCT de1.document_id) AS weight
                FROM document_entities de1
                JOIN document_entities de2 ON de2.document_id = de1.document_id
                JOIN documents d ON d.id = de1.document_id
                JOIN crawl_targets ct ON ct.id = d.target_id
                WHERE de1.canonical_name = %s AND de2.canonical_name = ANY(%s)
                    AND {where_clause}
                GROUP BY de2.canonical_name
                """,
                [resolved_focus, resolved_focus, others, *params],
            ).fetchall()
            return {"nodes": nodes, "edges": edges, "focus_canonical_name": resolved_focus}

        edges = conn.execute(
            f"""
            SELECT de1.canonical_name AS source, de2.canonical_name AS target,
                   COUNT(DISTINCT de1.document_id) AS weight
            FROM document_entities de1
            JOIN document_entities de2
                ON de2.document_id = de1.document_id AND de2.canonical_name > de1.canonical_name
            JOIN documents d ON d.id = de1.document_id
            JOIN crawl_targets ct ON ct.id = d.target_id
            WHERE de1.canonical_name = ANY(%s) AND de2.canonical_name = ANY(%s) AND {where_clause}
            GROUP BY de1.canonical_name, de2.canonical_name
            """,
            [own_names, own_names, *params],
        ).fetchall()

    return {"nodes": nodes, "edges": edges}
