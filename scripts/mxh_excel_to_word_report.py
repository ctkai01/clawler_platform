from __future__ import annotations

import argparse
import logging
from collections import defaultdict
from datetime import date

import openpyxl

from platform_app.reporting.word_report import build_daily_word_report_bytes

logger = logging.getLogger(__name__)

_SENTIMENT_LABEL = {"1": "positive", "0": "neutral", "-1": "negative"}


def _load_rows(xlsx_path: str, sheet_name: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    headers = [(c.value or "").strip() for c in ws[2]]
    rows = []
    for values in ws.iter_rows(min_row=3, values_only=True):
        row = dict(zip(headers, values))
        if (row.get("Tags") or "").strip() == "Spam":
            continue
        sentiment = _SENTIMENT_LABEL.get(str(row.get("Sentiment")).strip())
        if sentiment is None:
            continue
        reaction = row.get("Reaction") or 0
        comments = row.get("Comments") or 0
        shares = row.get("Shares") or 0
        title = (row.get("Title") or "").strip()
        intro = (row.get("Intro") or "").strip()
        if not title:
            title = intro[:80] + "…" if len(intro) > 80 else intro
        source_type = row.get("SourceType") or "Khác"
        source_name = (row.get("SourceName") or "").strip()
        url = row.get("URL") or row.get("ShareLink") or ""
        rows.append(
            {
                "sentiment": sentiment,
                "topic": row.get("TopicName") or "Các nội dung khác",
                "reaction": reaction,
                "comments": comments,
                "shares": shares,
                "engagement_total": reaction + comments,
                "title": title or "(không có tiêu đề)",
                "url": url,
                # The source tool exports one row per (post, matched-topic)
                "post_key": url or f"__no_url_{len(rows)}__",
                "channel_label": f"{source_type}: {source_name}" if source_name else source_type,
                "author": None,
            }
        )
    return rows


def _dedupe_posts(rows: list[dict]) -> list[dict]:
    """The same physical post is exported once per topic it matches (verified
    against the source file: identical post ID/URL/engagement repeated under
    '# MNP', '# DỊCH VỤ', '# MẠNG LƯỚI' for one post) — fine for the per-topic
    breakdown table, but KPI totals/sentiment counts/post lists must count
    each physical post once, keyed by URL (its only stable identity — the
    file's own ID/FBId columns are '0' for several rows)."""
    seen: dict[str, dict] = {}
    for r in rows:
        seen.setdefault(r["post_key"], r)
    return list(seen.values())


def _build_report(rows: list[dict], unique_rows: list[dict]) -> dict:
    # Topic breakdown intentionally uses `rows` (one entry per post/topic
    # match, duplicates included) — a post genuinely tagged under 3 topics
    # should count toward all 3. KPI/sentiment totals use `unique_rows` so
    # that same post isn't tripled just because it matched 3 topics.
    topic_stats: dict[str, dict[str, int]] = defaultdict(lambda: {"posts": 0, "comments": 0, "total_engagement": 0})
    topic_sentiment: dict[str, dict[str, int]] = defaultdict(lambda: {"positive": 0, "neutral": 0, "negative": 0})
    for r in rows:
        t = topic_stats[r["topic"]]
        t["posts"] += 1
        t["comments"] += r["comments"]
        t["total_engagement"] += r["engagement_total"]
        topic_sentiment[r["topic"]][r["sentiment"]] += 1

    topics_by_size = sorted(topic_stats.items(), key=lambda kv: -kv[1]["posts"])
    keyword_topic_detail = [{"topic": topic, **stats} for topic, stats in topics_by_size]
    topic_sentiment_rows = [{"topic": topic, **topic_sentiment[topic]} for topic, _ in topics_by_size]

    sentiment_counts = defaultdict(int)
    total_comments = total_reactions = total_shares = 0
    for r in unique_rows:
        sentiment_counts[r["sentiment"]] += 1
        total_comments += r["comments"]
        total_reactions += r["reaction"]
        total_shares += r["shares"]

    report = {
        "total_posts": len(unique_rows),
        "total_comments": total_comments,
        "total_reactions": total_reactions,
        "total_shares": total_shares,
        "sentiment_positive": sentiment_counts["positive"],
        "sentiment_neutral": sentiment_counts["neutral"],
        "sentiment_negative": sentiment_counts["negative"],
        "keyword_topic_detail": keyword_topic_detail,
    }
    return report, topic_sentiment_rows


def build_report_bytes(xlsx_path: str, sheet_name: str, org_name: str, period_label: str, report_date: date) -> bytes:
    rows = _load_rows(xlsx_path, sheet_name)
    unique_rows = _dedupe_posts(rows)
    report, topic_sentiment_rows = _build_report(rows, unique_rows)

    negative_posts_all = sorted(
        (r for r in unique_rows if r["sentiment"] == "negative"), key=lambda r: -r["engagement_total"]
    )
    positive_posts_all = sorted(
        (r for r in unique_rows if r["sentiment"] == "positive"), key=lambda r: -r["engagement_total"]
    )

    return build_daily_word_report_bytes(
        org_name=org_name,
        report_date=report_date,
        report=report,
        topic_sentiment_rows=topic_sentiment_rows,
        negative_posts=negative_posts_all[:5],
        positive_posts=positive_posts_all[:5],
        negative_total=len(negative_posts_all),
        positive_total=len(positive_posts_all),
        period_label=period_label,
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Sinh báo cáo Word từ file Excel dữ liệu MXH (export từ công cụ ngoài)")
    parser.add_argument("xlsx_path")
    parser.add_argument("output_path")
    parser.add_argument("--sheet", default="DS Noi Dung")
    parser.add_argument("--org-name", default="MobiFone")
    parser.add_argument("--period-label", required=True, help='VD: "TUẦN 07/07/2026 - 14/07/2026"')
    args = parser.parse_args()

    content = build_report_bytes(
        args.xlsx_path, args.sheet, args.org_name, args.period_label, date.today()
    )
    with open(args.output_path, "wb") as f:
        f.write(content)
    logger.info("Đã ghi báo cáo: %s (%d bytes)", args.output_path, len(content))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
