from __future__ import annotations

import argparse
import io
import logging
import re

import openpyxl
from docx import Document

from platform_app.db.pool import get_pool
from platform_app.pipeline.event_report import REACH_TIER_BY_SOURCE, _DEFAULT_REACH_TIER, generate_overview_narrative
from platform_app.pipeline.text_normalize import fold
from platform_app.reporting.event_word_report import build_event_weekly_word_report_bytes

logger = logging.getLogger(__name__)

_SENTIMENT_LABEL = {"1": "positive", "0": "neutral", "-1": "negative"}
# Same normalization the platform's own DB-driven event report uses
# (platform_app/pipeline/event_report.py's _COMPETITOR_BRAND_MAP) — kept in
# sync manually since this script reads an external xlsx export instead of
# organization_keywords.
_COMPETITOR_TERMS = {
    "viettel": "Viettel",
    "vinaphone": "VinaPhone",
    "vina phone": "VinaPhone",
    "vietnamobile": "Vietnamobile",
    "gmobile": "Gmobile",
}
# Bare "mobi"/"mbf" excluded — too generic (matches unrelated English text
# like a "MOBI PHONE REPAIR" shop bio), same fix and same reasoning as
# mxh_excel_to_word_report.py's _brand_patterns.
_TOO_GENERIC_BRAND_TERMS = {"mobi", "mbf"}
# This is specifically the "5G" event report — matching platform_app/pipeline
# /event_report.py's EVENT_DEFINITIONS["5g_mobifone"]["match_terms"] (a row
# must mention BOTH a brand AND the event term). Brand-only matching let a
# lot of non-5G MobiFone news through once a general (non-5G-scoped) xlsx
# export got merged in alongside the dedicated 5G export.
_EVENT_PATTERN = re.compile(r"\b5g\b", re.IGNORECASE)


def _brand_patterns(org_name: str) -> list[re.Pattern]:
    with get_pool().connection() as conn:
        rows = conn.execute(
            """
            SELECT kc.term FROM organization_keywords ok
            JOIN keywords_catalog kc ON kc.id = ok.keyword_id
            JOIN organizations o ON o.id = ok.organization_id
            WHERE o.name ILIKE %s AND kc.category = 'brand' AND kc.is_active
            """,
            (org_name,),
        ).fetchall()
    return [
        re.compile(r"\b" + re.escape(fold(row["term"])) + r"\b", re.IGNORECASE)
        for row in rows
        if fold(row["term"]) not in _TOO_GENERIC_BRAND_TERMS
    ]


def _load_rows(xlsx_path: str, sheet_name: str, brand_patterns: list[re.Pattern]) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    headers = [(c.value or "").strip() for c in ws[2]]
    rows = []
    for values in ws.iter_rows(min_row=3, values_only=True):
        row = dict(zip(headers, values))
        if (row.get("Tags") or "").strip() == "Spam":
            continue
        sentiment = _SENTIMENT_LABEL.get(str(row.get("Sentiment")).strip())
        if sentiment is None:
            continue

        title = (row.get("Title") or "").strip()
        intro = (row.get("Intro") or "").strip()
        display_title = title or (intro[:80] + "…" if len(intro) > 80 else intro) or "(không có tiêu đề)"
        text_lower = f"{title} {intro}".lower()
        text_folded = fold(f"{title} {intro}")

        if not _EVENT_PATTERN.search(text_folded):
            continue

        # Mirrors run_event_match's matched_brands set — a row can belong to
        # several brands at once (e.g. an article comparing MobiFone and
        # Viettel), and rows matching none are dropped entirely, same as the
        # DB-driven path (`if not matched_brands: continue`).
        brands = set()
        if any(p.search(text_folded) for p in brand_patterns):
            brands.add("MobiFone")
        for term, brand in _COMPETITOR_TERMS.items():
            if term in text_lower:
                brands.add(brand)
        if not brands:
            continue

        source_type = (row.get("SourceType") or "Khác").strip()
        source_name = (row.get("SourceName") or "").strip()
        url = row.get("URL") or row.get("ShareLink") or ""
        reaction = row.get("Reaction") or 0
        comments = row.get("Comments") or 0
        shares = row.get("Shares") or 0

        rows.append(
            {
                "brands": brands,
                "sentiment": sentiment,
                "title": display_title,
                "content": intro or title,
                "url": url,
                "post_key": url or f"__no_url_{len(rows)}__",
                "source_type": source_type,
                "source_name": source_name,
                "engagement_total": reaction + comments + shares,
                "is_news": source_type == "News",
            }
        )
    return rows


def _dedupe(rows: list[dict]) -> list[dict]:
    """The source tool exports one row per (post, matched-topic) — the same
    physical post can also legitimately appear twice here because it
    mentions two different brands in the SAME row (e.g. MobiFone vs Viettel
    comparison), so brand sets are merged rather than just keeping the
    first-seen copy."""
    seen: dict[str, dict] = {}
    for r in rows:
        existing = seen.get(r["post_key"])
        if existing is None:
            seen[r["post_key"]] = r
        else:
            existing["brands"] |= r["brands"]
    return list(seen.values())


def _quantile_cutoffs(values: list[int]) -> tuple[int, int]:
    """No LLM-assessed impact_level available for an external xlsx export —
    approximate it from this batch's own engagement distribution (top decile
    = Cao, next 30% = Trung bình, rest = Thấp) instead, per explicit user
    choice over paying for ~600 real OpenAI calls."""
    vals = sorted(v for v in values if v > 0)
    if not vals:
        return (1, 1)

    def q(p: float) -> int:
        return vals[min(len(vals) - 1, int(len(vals) * p))]

    return (q(0.90), q(0.60))


def _impact_level(value: int, cutoffs: tuple[int, int]) -> str:
    hi, mid = cutoffs
    if value >= hi:
        return "Cao"
    if value >= mid:
        return "Trung bình"
    return "Thấp"


def _sentiment_counts(matches: list[dict]) -> dict[str, int]:
    counts = {"positive": 0, "neutral": 0, "negative": 0}
    for m in matches:
        counts[m["sentiment"]] += 1
    return counts


def build_report_bytes(xlsx_paths: list[str], sheet_name: str, org_name: str, period_label: str) -> bytes:
    brand_patterns = _brand_patterns(org_name)
    # Same post can appear in more than one source export (e.g. a 5G-topic
    # export and a broader MobiFone export both crawled it) — merge first,
    # then dedupe by post_key across ALL files combined, not per file.
    rows = _dedupe([r for xlsx_path in xlsx_paths for r in _load_rows(xlsx_path, sheet_name, brand_patterns)])
    news_rows = [r for r in rows if r["is_news"]]
    social_rows = [r for r in rows if not r["is_news"]]
    news_cutoffs = _quantile_cutoffs([r["engagement_total"] for r in news_rows])
    social_cutoffs = _quantile_cutoffs([r["engagement_total"] for r in social_rows])

    def to_match(r: dict, brand: str, *, is_news: bool) -> dict:
        return {
            "brand": brand,
            "topic": r["title"],
            "content": r["content"],
            "url": r["url"],
            "target_name": r["source_name"] or r["source_type"],
            "platform_type": "news" if is_news else "social",
            "engagement_total": r["engagement_total"],
            "sentiment": r["sentiment"],
            "impact_level": _impact_level(r["engagement_total"], news_cutoffs if is_news else social_cutoffs),
            "handling_status": "chua_xu_ly",
            "reach_tier": REACH_TIER_BY_SOURCE.get(r["source_name"], _DEFAULT_REACH_TIER) if is_news else None,
        }

    news_matches = [to_match(r, b, is_news=True) for r in news_rows for b in r["brands"]]
    social_matches_all = [to_match(r, b, is_news=False) for r in social_rows for b in r["brands"]]

    mobifone_news = [m for m in news_matches if m["brand"] == org_name]
    competitor_brands = sorted({b for r in rows for b in r["brands"] if b != org_name})
    competitor_news: dict[str, list[dict]] = {b: [] for b in competitor_brands}
    for m in news_matches:
        if m["brand"] != org_name:
            competitor_news.setdefault(m["brand"], []).append(m)
    social_mobifone = [m for m in social_matches_all if m["brand"] == org_name]

    # No equivalent last-week export in this same "DS Noi Dung" xlsx format
    # exists to compare against — per explicit user choice, previous-week
    # numbers are left at 0 and the "so sánh" column is overwritten to "—"
    # below (a real 0 baseline would otherwise render a misleading "100%").
    comparison = {
        "yesterday_label": "Không có dữ liệu",
        "today_label": period_label,
        "news": {
            "yesterday_total": 0,
            "today_total": len(news_matches),
            "yesterday_sentiment": {"positive": 0, "neutral": 0, "negative": 0},
            "today_sentiment": _sentiment_counts(news_matches),
        },
        "social": {
            "yesterday_total": 0,
            "today_total": len(social_matches_all),
            "yesterday_sentiment": {"positive": 0, "neutral": 0, "negative": 0},
            "today_sentiment": _sentiment_counts(social_matches_all),
        },
    }

    brand_counts = {
        b: _sentiment_counts([m for m in news_matches + social_matches_all if m["brand"] == b])
        for b in [org_name, *competitor_brands]
    }

    overview_narrative = generate_overview_narrative(org_name, news_matches)

    content = build_event_weekly_word_report_bytes(
        org_name=org_name,
        event_label="5G",
        period_label=period_label,
        comparison=comparison,
        overview_narrative=overview_narrative,
        mobifone_news=mobifone_news,
        competitor_news=competitor_news,
        social_matches=social_mobifone,
        brand_counts=brand_counts,
    )

    doc = Document(io.BytesIO(content))
    # tables[0] is the brand-vs-brand sentiment summary (added first by
    # build_event_weekly_word_report_bytes); tables[1] is the yesterday/today
    # comparison table (9 rows x 5 cols) whose last column we're overriding.
    comparison_table = doc.tables[1]
    for row in comparison_table.rows[1:]:
        row.cells[-1].text = "—"
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Sinh báo cáo Word 'BÁO CÁO ONLINE MẠNG XÃ HỘI VÀ BÁO CHÍ VỀ 5G MOBIFONE' từ file Excel DS Noi Dung"
    )
    parser.add_argument("xlsx_paths", nargs="+", help="1 hoặc nhiều file .xlsx (gộp dữ liệu, dedupe theo URL)")
    parser.add_argument("output_path")
    parser.add_argument("--sheet", default="DS Noi Dung")
    parser.add_argument("--org-name", default="MobiFone")
    parser.add_argument("--period-label", required=True, help='VD: "13/07 - 19/07"')
    args = parser.parse_args()

    content = build_report_bytes(args.xlsx_paths, args.sheet, args.org_name, args.period_label)
    with open(args.output_path, "wb") as f:
        f.write(content)
    logger.info("Đã ghi báo cáo: %s (%d bytes)", args.output_path, len(content))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
