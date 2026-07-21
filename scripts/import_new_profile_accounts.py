from __future__ import annotations

import json
import logging
from pathlib import Path

import openpyxl
from psycopg.types.json import Jsonb

from platform_app.db.pool import get_pool

logger = logging.getLogger(__name__)

SESSION_DIR = Path("/home/ctkai/Documents/facebook_profile")
XLSX_PATH = SESSION_DIR / "Facebook_Accounts_Structured.xlsx"


def _load_credentials() -> dict[str, dict]:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Sheet1"]
    headers = [c.value for c in ws[1]]
    idx_uid = headers.index("UID")
    idx_2fa = headers.index("2FA / Secret")
    idx_pw = headers.index("Password")

    creds = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        uid = str(row[idx_uid])
        creds[uid] = {"two_fa_secret": row[idx_2fa], "password": row[idx_pw]}
    return creds


def main() -> int:
    creds = _load_credentials()
    session_paths = sorted(SESSION_DIR.glob("*_storage_state.json"))

    imported = []
    with get_pool().connection() as conn:
        for path in session_paths:
            uid = path.stem.replace("_storage_state", "")
            session_data = json.loads(path.read_text(encoding="utf-8"))
            cred = creds.get(uid, {})
            conn.execute(
                """
                INSERT INTO fb_accounts (id, session_data, password, two_fa_secret, status, fail_count)
                VALUES (%s, %s, %s, %s, 'LIVE', 0)
                ON CONFLICT (id) DO UPDATE SET
                    session_data = EXCLUDED.session_data,
                    password = EXCLUDED.password,
                    two_fa_secret = EXCLUDED.two_fa_secret,
                    status = 'LIVE',
                    fail_count = 0,
                    updated_at = now()
                """,
                (uid, Jsonb(session_data), cred.get("password"), cred.get("two_fa_secret")),
            )
            imported.append(uid)

    logger.info("Đã import %d account: %s", len(imported), ", ".join(imported))
    return 0


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    raise SystemExit(main())
